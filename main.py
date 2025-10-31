import os
import pickle
import io
import re
import unicodedata
from typing import Iterable, Tuple, Optional, Union, Dict, List
from dotenv import load_dotenv

from telebot import TeleBot, types
from openai import OpenAI
import numpy as np
import pandas as pd
from pandas import ExcelWriter

from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

load_dotenv()

# ===================== CONFIG =====================
BOT_TOKEN = os.getenv("BOT_TOKEN")
MAIN_SHEET_NAME = "BUDG"  # read this sheet from the main file
ALLOWED_MAIN_COLUMNS = ["Назва Офферу", "ГЕО", "Загальні витрати"]
ADDITIONAL_REQUIRED_COLS = ["Країна", "Сума депозитів"]

# Flexible synonyms (users can rename columns — we’ll still find them)
GEO_COL_CANDIDATES = ["ГЕО", "GEO", "Geo"]
OFFER_COL_CANDIDATES = ["Назва Офферу", "Оффер", "Offer", "Назва оффера", "Назва офферу"]

DEFAULT_PAIRS: list[tuple[str, str]] = [
    ("Аргентина", "TRAFCODE"),
    ("Бенін", "TRAFCODEx"),
]

OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-5")
OPENAI_MAX_CHARS = 60_000  # безпечний ліміт для одного запиту
OPENAI_OUTPUT_COLUMN = "Total spend"  # колонка, яку модель має заповнити/перерахувати

bot = TeleBot(BOT_TOKEN, parse_mode="HTML")

H_THRESH = 9.0  # H <= 9 is "good"
L_THRESH = 39.99  # L > 39.99 is "good" (strict >)
CPA_CAP = 11.0
EPS = 1e-12
EPS_YEL = 1e-6

# CPA Target defaults and overrides
CPA_DEFAULT_TARGET = 8
CPA_OVERRIDES: Dict[str, float] = {
    "Аргентина": 20,
    "Болівія": 15,
    "Венесуела": 5,
    "Габон": 7,
    "Гана": 5,
    "Еквадор": 15,
    "Йорданія": 40,
    "Ірак": 40,
    "Казахстан": 30,
    "Колумбія": 11,
    "Малайзія": 40,
    "Парагвай": 15,
    "Пакистан": 15,
    "Перу": 12,
    "Таїланд": 22,
    "Уругвай": 12,
    "Філіппіни": 10,
}


# ===================== STATE =====================
class UserState:
    def __init__(self):
        self.alloc_mode = None
        self.phase = "WAIT_MAIN"  # WAIT_MAIN -> WAIT_ADDITIONAL, plus allocate flow
        self.main_agg_df: Optional[pd.DataFrame] = None
        self.offers: List[str] = []
        self.current_offer_index: int = 0
        self.offer_deposits: Dict[str, Dict[str, Dict[str, float]]] = {}

        # country maps
        self.country_map_uk_to_en = build_country_map_uk_to_en()
        self.country_map_ru_to_en = build_country_map_ru_to_en()
        self.country_canon = build_country_canonical()

        # --- allocate flow state ---
        self.alloc_df: Optional[pd.DataFrame] = None
        self.alloc_budget: Optional[float] = None
        self.alloc_pairs: Optional[list[tuple[str, str]]] = None
        # NEW: optional sets to control filtering and Total+% overrides
        self.current_pairs: Optional[set[tuple[str, str]]] = None
        self.plus35_pairs: Optional[set[tuple[str, str]]] = None


user_states: Dict[int, UserState] = {}

# ===== ACCESS CONTROL =====
# Retrieve and parse ALLOWED_USER_IDS
allowed_user_ids = os.getenv("ALLOWED_USER_IDS")
ALLOWED_USER_IDS = {uid.strip() for uid in allowed_user_ids.split(",") if uid.strip()}

print("Loaded user IDs:", ALLOWED_USER_IDS)

USERS_FILE = os.path.join("data", "telegram_users.txt")


def save_user_if_new(user: types.User, path: str = USERS_FILE) -> None:
    """
    Зберігає рядок: <id>\t<first_name>\t<last_name>\t<username>
    Додає лише, якщо user.id ще відсутній у файлі.
    """
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        user_id = str(user.id)
        first_name = user.first_name or ""
        last_name = user.last_name or ""
        username = user.username or ""

        # Відкриваємо у режимі a+ (створить файл, якщо його немає),
        # читаємо для перевірки і, за потреби, дописуємо.
        with open(path, "a+", encoding="utf-8") as f:
            f.seek(0)
            exists = any(line.split("\t", 1)[0] == user_id for line in f)
            if not exists:
                f.write(f"{user_id}\t{first_name}\t{last_name}\t{username}\n")
    except Exception as e:
        # Не ламаємо логіку бота, просто логнемо (за наявності logger)
        try:
            print(f"Failed to save user to {path}: {e}")
        except:
            pass


def _deny_access_message():
    return (
        "⛔ <b>Доступ заборонено.</b>\n"
        "Якщо вам потрібен доступ — зверніться до адміністратора бота."
    )


def _is_allowed_user(user_id: int) -> bool:
    return str(user_id) in ALLOWED_USER_IDS


# Для message-хендлерів
def require_access(handler_func):
    def wrapper(message, *args, **kwargs):
        user_id = getattr(message.from_user, "id", None)
        if user_id is None or not _is_allowed_user(user_id):
            bot.reply_to(message, _deny_access_message())
            return
        return handler_func(message, *args, **kwargs)

    return wrapper


# Для callback-query (інлайн-кнопки)
def require_access_cb(handler_func):
    def wrapper(call, *args, **kwargs):
        user_id = getattr(call.from_user, "id", None)
        if user_id is None or not _is_allowed_user(user_id):
            try:
                bot.answer_callback_query(call.id, "⛔ Немає доступу.")
            except Exception:
                pass
            bot.send_message(call.message.chat.id, _deny_access_message())
            return
        return handler_func(call, *args, **kwargs)

    return wrapper


def _df_to_csv(df: pd.DataFrame) -> str:
    # Без індексу, максимально “плоско”
    bio = io.StringIO()
    df.to_csv(bio, index=False)
    return bio.getvalue()


def _split_df_by_size(df: pd.DataFrame, max_chars: int = OPENAI_MAX_CHARS) -> list[pd.DataFrame]:
    """
    Ділимо датафрейм на чанки, щоб CSV кожного не перевищував ліміт символів.
    """
    # груба оцінка середнього розміру рядка
    sample = min(len(df), 20)
    avg_row_len = len(_df_to_csv(df.head(sample))) / max(sample, 1)
    # запас: шапка + промпт
    rows_per_chunk = max(5, int((max_chars - 3000) / max(avg_row_len, 1)))
    chunks = []
    for i in range(0, len(df), rows_per_chunk):
        chunks.append(df.iloc[i:i + rows_per_chunk].copy())
    return chunks


def _csv_from_text(text: str) -> str:
    """
    Витягує CSV з відповіді (підтримка варіантів без код-блоків та з ```csv ... ```).
    """
    t = text.strip()
    if "```" in t:
        # намагаємося знайти fenced block
        parts = t.split("```")
        # шукаємо блок з csv або перший код-блок
        best = None
        for i in range(1, len(parts), 2):
            block = parts[i]
            if block.lstrip().lower().startswith("csv"):
                best = block.split("\n", 1)[1] if "\n" in block else ""
                break
        if best is None:
            # беремо перший код-блок як fallback
            best = parts[1]
        return best.strip()
    return t


SAVE_FILE = "saved_pairs.pkl"

# Global caches (used as single source of truth)
GLOBAL_CURRENT_PAIRS: set[tuple[str, str]] | None = None
GLOBAL_PLUS35_PAIRS: set[tuple[str, str]] | None = None


def save_pairs() -> None:
    """Persist global pairs once for the whole server."""
    data = {
        "current_pairs": GLOBAL_CURRENT_PAIRS,
        "plus35_pairs": GLOBAL_PLUS35_PAIRS,
    }
    with open(SAVE_FILE, "wb") as f:
        pickle.dump(data, f)


def load_pairs() -> None:
    """Load global pairs into GLOBAL_* variables (no per-user binding)."""
    global GLOBAL_CURRENT_PAIRS, GLOBAL_PLUS35_PAIRS
    if not os.path.exists(SAVE_FILE):
        GLOBAL_CURRENT_PAIRS, GLOBAL_PLUS35_PAIRS = None, None
        return
    with open(SAVE_FILE, "rb") as f:
        data = pickle.load(f) or {}
    GLOBAL_CURRENT_PAIRS = data.get("current_pairs")
    GLOBAL_PLUS35_PAIRS = data.get("plus35_pairs")


load_pairs()


# ===================== NORMALIZATION (countries) =====================
def normalize_text(s: str) -> str:
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def build_country_map_uk_to_en() -> Dict[str, str]:
    m = {
        "Аргентина": "Argentina",
        "Бенін": "Benin",
        "Буркіна-Фасо": "Burkina Faso",
        "Венесуела": "Venezuela",
        "Габон": "Gabon",
        "Гаїті": "Haiti",
        "Гана": "Ghana",

        # ❗ уточнення: у XLSX зустрічається "Guinea-Conakry"
        # щоб зв’язати з дод. таблицями, мапимо укр "Гвінея" саме на "Guinea-Conakry"
        # (далі в каноні зведемо це до "Guinea")
        "Гвінея": "Guinea-Conakry",

        # ---- DRC/ROC ----
        "Демократична Республіка Конґо": "Congo (Kinshasa)",
        "ДР Конго": "Congo (Kinshasa)",
        "Конго (Кіншаса)": "Congo (Kinshasa)",
        "Республіка Конго": "Congo (Brazzaville)",
        "Конго-Браззавіль": "Congo (Brazzaville)",
        # якщо просто "Конго" — за замовчуванням DRC
        "Конго": "Congo (Brazzaville)",

        "Камерун": "Cameroon",

        # Кот-д’Івуар — одразу кілька апострофних варіантів
        "Кот-д'Івуар": "Cote d'Ivoire",
        "Кот-д’Івуар": "Cote d'Ivoire",
        "Кот д’Івуар": "Cote d'Ivoire",

        "Кенія": "Kenya",
        "Сенегал": "Senegal",
        "Сьєрра-Леоне": "Sierra Leone",
        "Танзанія": "Tanzania",
        "Того": "Togo",
        "Уганда": "Uganda",
        "Замбія": "Zambia",
        "Ефіопія": "Ethiopia",
        "Нігер": "Niger",
        "Нігерія": "Nigeria",
        "Малі": "Mali",
        "Пакистан": "Pakistan",
        "Казахстан": "Kazakhstan",
        "Іспанія": "Spain",
        "Франція": "France",
        "Італія": "Italy",
        "Португалія": "Portugal",
        "Домініканська Республіка": "Dominican Republic",
        "Канада": "Canada",
        "Філіппіни": "Philippines",

        # 🔹 додано з вашого списку «missing»
        "Болівія": "Bolivia",
        "Еквадор": "Ecuador",
        "Колумбія": "Colombia",
        "Парагвай": "Paraguay",
        "Перу": "Peru",
    }
    return {normalize_text(k): v for k, v in m.items()}


def build_country_map_ru_to_en() -> Dict[str, str]:
    m = {
        "Аргентина": "Argentina",
        "Бенин": "Benin",
        "Буркина-Фасо": "Burkina Faso",
        "Венесуэла": "Venezuela",
        "Габон": "Gabon",
        "Гаити": "Haiti",
        "Гана": "Ghana",

        # Как и в UA-карте: "Гвинея" → "Guinea-Conakry" (далее канонизируем в "Guinea")
        "Гвинея": "Guinea-Conakry",

        # ---- DRC/ROC ----
        "Демократическая Республика Конго": "Congo (Kinshasa)",
        "ДР Конго": "Congo (Kinshasa)",
        "Конго (Киншаса)": "Congo (Kinshasa)",
        "Республика Конго": "Congo (Brazzaville)",
        "Конго-Браззавиль": "Congo (Brazzaville)",
        # если просто "Конго" — как в UA-карте используем Brazzaville
        "Конго": "Congo (Brazzaville)",

        "Камерун": "Cameroon",

        # Кот-д’Ивуар — разные апострофы/пробелы
        "Кот-д'Ивуар": "Cote d'Ivoire",
        "Кот-д’Ивуар": "Cote d'Ivoire",
        "Кот д’Ивуар": "Cote d'Ivoire",

        "Кения": "Kenya",
        "Сенегал": "Senegal",
        "Сьерра-Леоне": "Sierra Leone",
        "Танзания": "Tanzania",
        "Того": "Togo",
        "Уганда": "Uganda",
        "Замбия": "Zambia",
        "Эфиопия": "Ethiopia",
        "Нигер": "Niger",
        "Нигерия": "Nigeria",
        "Мали": "Mali",
        "Пакистан": "Pakistan",
        "Казахстан": "Kazakhstan",
        "Испания": "Spain",
        "Франция": "France",
        "Италия": "Italy",
        "Португалия": "Portugal",
        "Доминиканская Республика": "Dominican Republic",
        "Канада": "Canada",
        "Филиппины": "Philippines",

        # Латам (как в UA-карте)
        "Боливия": "Bolivia",
        "Эквадор": "Ecuador",
        "Колумбия": "Colombia",
        "Парагвай": "Paraguay",
        "Перу": "Peru",
    }
    return {normalize_text(k): v for k, v in m.items()}


def build_country_canonical() -> Dict[str, str]:
    canon = {
        # самоканонічні EN
        "Argentina": "Argentina",
        "Benin": "Benin",
        "Burkina Faso": "Burkina Faso",
        "Gabon": "Gabon",
        "Haiti": "Haiti",
        "Ghana": "Ghana",
        "Guinea": "Guinea",
        "Congo (Kinshasa)": "Congo (Kinshasa)",
        "Congo (Brazzaville)": "Congo (Brazzaville)",
        "Cameroon": "Cameroon",
        "Cote d'Ivoire": "Cote d'Ivoire",
        "Kenya": "Kenya",
        "Senegal": "Senegal",
        "Sierra Leone": "Sierra Leone",
        "Tanzania": "Tanzania",
        "Togo": "Togo",
        "Uganda": "Uganda",
        "Venezuela": "Venezuela",
        "Zambia": "Zambia",
        "Ethiopia": "Ethiopia",
        "Niger": "Niger",
        "Nigeria": "Nigeria",
        "Mali": "Mali",
        "Pakistan": "Pakistan",
        "Kazakhstan": "Kazakhstan",
        "Spain": "Spain",
        "France": "France",
        "Italy": "Italy",
        "Portugal": "Portugal",
        "Dominican Republic": "Dominican Republic",
        "Canada": "Canada",
        "Philippines": "Philippines",

        # 🔹 додано з вашого списку «missing»
        "Bolivia": "Bolivia",
        "Ecuador": "Ecuador",
        "Colombia": "Colombia",
        "Paraguay": "Paraguay",
        "Peru": "Peru",

        # синоніми/варіанти написання → канон
        # Cote d'Ivoire
        "Cote DIvoire": "Cote d'Ivoire",
        "Cote dIvoire": "Cote d'Ivoire",
        "Cote D Ivoire": "Cote d'Ivoire",
        "Cote d’ivoire": "Cote d'Ivoire",
        "Côte d’Ivoire": "Cote d'Ivoire",
        "Ivory Coast": "Cote d'Ivoire",

        # Guinea-Conakry → Guinea
        "Guinea-Conakry": "Guinea",
        "Guinea Conakry": "Guinea",
        "Guinea, Conakry": "Guinea",

        # DRC/ROC варіанти
        "DRC": "Congo (Kinshasa)",
        "DR Congo": "Congo (Kinshasa)",
        "Congo (DRC)": "Congo (Kinshasa)",
        "Democratic Republic of the Congo": "Congo (Kinshasa)",
        "Democratic Republic of Congo": "Congo (Kinshasa)",
        "Congo-Kinshasa": "Congo (Kinshasa)",

        "Republic of the Congo": "Congo (Brazzaville)",
        "Congo Republic": "Congo (Brazzaville)",
        "Congo-Brazzaville": "Congo (Brazzaville)",

        # UA → EN канон (на випадок, якщо десь просочиться укр у дод. таблицях)
        "кот-д'івуар": "Cote d'Ivoire",
        "кот-д’івуар": "Cote d'Ivoire",
        "кот д’івуар": "Cote d'Ivoire",
        "гвінея": "Guinea",
        "болівія": "Bolivia",
        "еквадор": "Ecuador",
        "колумбія": "Colombia",
        "парагвай": "Paraguay",
        "перу": "Peru",
    }
    return {normalize_text(k): v for k, v in canon.items()}


def to_canonical_en(
        country: str,
        uk_to_en: Dict[str, str],
        canonical: Dict[str, str],
        ru_to_en: Optional[Dict[str, str]] = None,
) -> str:
    key = normalize_text(country)

    # 1) UA → EN
    if key in uk_to_en:
        mapped = uk_to_en[key]
        return canonical.get(normalize_text(mapped), mapped)

    # 2) RU → EN (резерв, якщо не знайшли в UA)
    if ru_to_en and key in ru_to_en:
        mapped = ru_to_en[key]
        return canonical.get(normalize_text(mapped), mapped)

    # 3) Уже EN/варіант — канонізуємо
    if key in canonical:
        return canonical[key]

    # 4) Спецвипадки
    if key in {"конго", "congo"}:
        return "Congo (Kinshasa)"

    # 5) Як є (не впізнали)
    return country


def cpa_target_for_geo(geo: Optional[str]) -> float:
    try:
        key = str(geo).strip()
    except Exception:
        key = ""
    return CPA_OVERRIDES.get(key, CPA_DEFAULT_TARGET)


# ===================== FLEXIBLE HEADER / COLUMN MATCH =====================

def detect_header_row(df: pd.DataFrame, required: List[str], max_scan: int = 60,
                      scan_rows: Optional[int] = None) -> int:
    # accept either max_scan or scan_rows for backward compatibility
    if scan_rows is not None:
        max_scan = scan_rows
    req_norm = [normalize_text(r) for r in required]
    for i in range(min(max_scan, len(df))):
        row_vals = [normalize_text(x) for x in list(df.iloc[i].values)]
        if all(any(h == v for v in row_vals) for h in req_norm):
            return i
    return 0


def match_columns(actual_cols, required_labels: List[str]) -> Optional[Dict[str, str]]:
    norm_actual = {normalize_text(str(c)): str(c) for c in actual_cols}
    out = {}
    for need in required_labels:
        key = normalize_text(need)
        if key in norm_actual:
            out[need] = norm_actual[key]
            continue
        # punctuation-insensitive
        found = None
        for k, v in norm_actual.items():
            if normalize_text(re.sub(r"[^\w\s'’]", "", k)) == normalize_text(re.sub(r"[^\w\s'’]", "", key)):
                found = v
                break
        if not found:
            return None
        out[need] = found
    return out


# ===================== FILE READERS =====================
# ADD
def read_excel_robust(file_bytes: bytes, sheet_name: str, header: int = 0) -> pd.DataFrame:
    """
    Robust Excel reader with multiple fallback strategies.
    Filters to the current month:
      - Prefer column 'Місяць' (numeric month: 1..12).
      - Fallback to column 'Дата' (dd/mm/YYYY).
    """
    bio = io.BytesIO(file_bytes)
    errors = []

    # Helper: filter to current month
    def filter_current_month(df: pd.DataFrame) -> pd.DataFrame:
        cur_month = datetime.now().month

        if "Місяць" in df.columns:
            # Accept strings like "09", numbers like 9.0, etc.
            month_series = pd.to_numeric(df["Місяць"], errors="coerce")
            return df[month_series == cur_month]

        return df

    # Strategy 1: openpyxl
    try:
        bio.seek(0)
        df = pd.read_excel(bio, sheet_name=sheet_name, header=header, engine="openpyxl")
        return filter_current_month(df)
    except Exception as e:
        errors.append(f"openpyxl: {e}")
        print(f"openpyxl failed: {e}")

    # Strategy 2: openpyxl without header, set manually
    try:
        bio.seek(0)
        df_raw = pd.read_excel(bio, sheet_name=sheet_name, header=None, engine="openpyxl")
        if header > 0:
            new_header = df_raw.iloc[header].tolist()
            df = df_raw.iloc[header + 1:].reset_index(drop=True)
            df.columns = new_header
        else:
            df = df_raw
        return filter_current_month(df)
    except Exception as e:
        errors.append(f"openpyxl manual header: {e}")
        print(f"Manual header setting failed: {e}")

    # Strategy 3: calamine
    try:
        bio.seek(0)
        df = pd.read_excel(bio, sheet_name=sheet_name, header=header, engine="calamine")
        return filter_current_month(df)
    except Exception as e:
        errors.append(f"calamine: {e}")
        print(f"calamine failed: {e}")

    # Strategy 4: xlrd
    try:
        bio.seek(0)
        df = pd.read_excel(bio, sheet_name=sheet_name, header=header, engine="xlrd")
        return filter_current_month(df)
    except Exception as e:
        errors.append(f"xlrd: {e}")
        print(f"xlrd failed: {e}")

    raise ValueError(f"Could not read Excel file with any engine. Errors: {'; '.join(errors)}")


def load_pairs_table(
        pairs_source: Union[str, pd.DataFrame],
        *,
        geo_col: Optional[str] = None,
        offer_col: Optional[str] = None,
) -> list[tuple[str, str]]:
    """
    Load a user-submitted table of unique (ГЕО, Оффер) pairs.
    - pairs_source: path to .csv/.xlsx OR a pandas DataFrame
    - geo_col/offer_col: optional explicit column names (override auto-detect)
    Returns: list of (geo, offer) tuples, de-duplicated, stripped, non-empty.
    Raises: ValueError with a friendly message on problems.
    """
    # 1) Load
    if isinstance(pairs_source, pd.DataFrame):
        df = pairs_source.copy()
    elif isinstance(pairs_source, str):
        if pairs_source.lower().endswith(".csv"):
            df = pd.read_csv(pairs_source)
        elif pairs_source.lower().endswith((".xlsx", ".xls")):
            df = pd.read_excel(pairs_source)
        else:
            raise ValueError("Unsupported file type. Use .csv or .xlsx, or pass a DataFrame.")
    else:
        raise ValueError("pairs_source must be a DataFrame or a path to .csv/.xlsx")

    # 2) Find columns
    def pick_col(cands: Iterable[str]) -> Optional[str]:
        for c in cands:
            if c in df.columns:
                return c
        return None

    g_col = geo_col or pick_col(GEO_COL_CANDIDATES)
    o_col = offer_col or pick_col(OFFER_COL_CANDIDATES)

    if not g_col or not o_col:
        raise ValueError(
            "Table must contain columns for GEO and Offer. "
            f"Tried GEO in {GEO_COL_CANDIDATES}, Offer in {OFFER_COL_CANDIDATES}. "
            "You can also pass geo_col=... and offer_col=...."
        )

    # 3) Clean & validate rows
    sub = (
        df[[g_col, o_col]]
        .astype(str)
        .apply(lambda s: s.str.strip())
        .dropna()
        .replace({"": pd.NA})
        .dropna()
        .drop_duplicates()
    )

    if sub.empty:
        raise ValueError("Pairs table is empty after cleaning (no valid (ГЕО, Оффер) rows).")

    # 4) To list of tuples
    pairs: list[tuple[str, str]] = list(sub.itertuples(index=False, name=None))
    return pairs


def load_main_budg_table(file_bytes: bytes, filename: str = "uploaded") -> pd.DataFrame:
    df = None
    errors = []

    if filename.lower().endswith((".xlsx", ".xls", ".xlsm")):
        try:
            df = read_excel_robust(file_bytes, sheet_name="BUDG", header=1)
        except Exception as e1:
            errors.append(f"header=1: {e1}")
            # (your existing Excel fallback logic remains unchanged)
    else:
        # CSV support
        bio = io.BytesIO(file_bytes)
        try:
            df_raw = pd.read_csv(bio, header=None)
        except Exception:
            bio.seek(0)
            try:
                df_raw = pd.read_csv(bio, header=None, encoding="cp1251")
            except Exception:
                bio.seek(0)
                df_raw = pd.read_csv(bio, header=None, encoding="utf-8")

        # detect header row (similar to Excel logic)
        header_row = -1
        for i in range(min(10, len(df_raw))):
            row_values = [str(v).lower().strip() for v in df_raw.iloc[i].values]
            if any("назва" in val and "оффер" in val for val in row_values) and \
                    any("гео" in val for val in row_values) and \
                    any("витрат" in val for val in row_values):
                header_row = i
                break

        if header_row >= 0:
            new_header = df_raw.iloc[header_row].tolist()
            df = df_raw.iloc[header_row + 1:].reset_index(drop=True)
            df.columns = new_header
        else:
            df = df_raw

        # --- filter current month ---
        # cur_month = datetime.now().month
        cur_month = 10

        if "Місяць" in df.columns:
            df["Місяць"] = pd.to_numeric(df["Місяць"], errors="coerce")
            df = df[df["Місяць"] == cur_month]
        elif "Дата" in df.columns:
            df["Дата"] = pd.to_datetime(df["Дата"], format="%d/%m/%Y", errors="coerce")
            df = df[(df["Дата"].dt.month == cur_month) & (df["Дата"].dt.year == datetime.now().year)]

    if df is None:
        raise ValueError(f"Could not load BUDG sheet. Errors: {'; '.join(errors)}")

    # clean column names
    df.columns = [str(c).replace("\xa0", " ").replace("\u00A0", " ").strip() for c in df.columns]

    # validate and map columns
    colmap = match_columns(df.columns, ALLOWED_MAIN_COLUMNS)
    if not colmap:
        available = [str(c) for c in df.columns]
        required = ALLOWED_MAIN_COLUMNS
        raise ValueError(
            f"У BUDG мають бути колонки: {required}.\n"
            f"Доступні колонки: {available}\n"
            f"Перевір назви колонок у файлі."
        )

    df = df[[colmap["Назва Офферу"], colmap["ГЕО"], colmap["Загальні витрати"]]].copy()
    df.columns = ["Назва Офферу", "ГЕО", "Загальні витрати"]

    return df


def read_additional_table(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """
    Read additional table with improved error handling
    """
    bio = io.BytesIO(file_bytes)

    if filename.lower().endswith((".xlsx", ".xls", ".xlsm")):
        # Try multiple strategies for Excel files
        try:
            df_raw = pd.read_excel(bio, header=None, engine="openpyxl")
        except Exception:
            try:
                bio.seek(0)
                df_raw = pd.read_excel(bio, header=None, engine="calamine")
            except Exception:
                bio.seek(0)
                df_raw = pd.read_excel(bio, header=None, engine="xlrd")
    else:
        # CSV files
        try:
            df_raw = pd.read_csv(bio, header=None)
        except Exception:
            bio.seek(0)
            try:
                df_raw = pd.read_csv(bio, header=None, encoding="cp1251")
            except Exception:
                bio.seek(0)
                df_raw = pd.read_csv(bio, header=None, encoding="utf-8")

    # Find header row
    header_idx = detect_header_row(df_raw, [normalize_text(x) for x in ADDITIONAL_REQUIRED_COLS])

    # Validate header row exists
    if header_idx >= len(df_raw):
        raise ValueError(f"Не знайдено заголовків у файлі {filename}")

    headers = df_raw.iloc[header_idx].tolist()
    data = df_raw.iloc[header_idx + 1:].reset_index(drop=True)
    data.columns = headers

    # Clean column names
    data.columns = [str(c).replace("\xa0", " ").replace("\u00A0", " ").strip() for c in data.columns]

    col_map = match_columns(data.columns, ADDITIONAL_REQUIRED_COLS)
    if not col_map:
        available = [str(c) for c in data.columns]
        required = ADDITIONAL_REQUIRED_COLS
        raise ValueError(
            f"У файлі {filename} мають бути колонки: {required}.\n"
            f"Доступні колонки: {available}\n"
            f"Перевір назви колонок у файлі."
        )

    df = data[[col_map["Країна"], col_map["Сума депозитів"]]].copy()
    df.columns = ["Країна", "Сума депозитів"]
    return df


# ===================== HELPERS =====================

def inject_formulas_and_cf(
        ws,
        *,
        header_row: int = 1,
        first_data_row: int = 2,
        last_data_row: int | None = None,
):
    """
    Додає формули у колонки та CF-підсвітку.
    Очікується, що у шапці вже є принаймні колонки:
    - 'Total spend' (F), 'FTD qty' (E), 'Total Dep Amount' (K), 'My deposit amount' (L)
    Якщо інших колонок (Total+%, CPA, СP/Ч, Target 40/50%) немає — створимо.

    Формули:
      G: Total+%                 = Total spend * 1.3
      H: CPA                     = Total+% / FTD qty
      (кирилиця) 'СP/Ч'          = Total Dep Amount / FTD qty
      'C. profit Target 40%'     = Total+% * 0.4
      'C. profit Target 50%'     = Total+% * 0.5
      L: My deposit amount       = Total Dep Amount / Total+% * 100

    """

    if last_data_row is None:
        last_data_row = ws.max_row
    if last_data_row < first_data_row:
        return

    # --- Map header name -> column index (створюємо, якщо треба)
    headers = {ws.cell(row=header_row, column=c).value: c for c in range(1, ws.max_column + 1) if
               ws.cell(row=header_row, column=c).value}

    def ensure_col(name: str) -> int:
        if name in headers:
            return headers[name]
        col = ws.max_column + 1
        ws.cell(row=header_row, column=col, value=name)
        headers[name] = col
        return col

    col_idx = {
        "Total spend": ensure_col("Total spend"),
        "FTD qty": ensure_col("FTD qty"),
        "Total Dep Amount": ensure_col("Total Dep Amount"),
        "My deposit amount": ensure_col("My deposit amount"),
        "Total+%": ensure_col("Total+%"),
        "CPA": ensure_col("CPA"),
        "CPA Target": ensure_col("CPA Target"),
        "СP/Ч": ensure_col("СP/Ч"),  # перша літера — кирилична "С"
        "C. profit Target 40%": ensure_col("C. profit Target 40%"),
        "C. profit Target 50%": ensure_col("C. profit Target 50%"),
    }

    # У зручні змінні — кол. літери
    letter = {k: get_column_letter(v) for k, v in col_idx.items()}

    F = letter["Total spend"]
    E = letter["FTD qty"]
    K = letter["Total Dep Amount"]
    L = letter["My deposit amount"]
    G = letter["Total+%"]
    H = letter["CPA"]
    I = letter["CPA Target"]
    CPCH = letter["СP/Ч"]
    C40 = letter["C. profit Target 40%"]
    C50 = letter["C. profit Target 50%"]

    # --- Прописуємо формули по рядках
    mult_col_letter = None
    for cell in ws[header_row]:
        if str(cell.value).strip().lower() == 'multiplier':
            mult_col_letter = cell.column_letter
            break
    for r in range(first_data_row, last_data_row + 1):
        if mult_col_letter:
            ws[f"{G}{r}"] = f"={F}{r}*{mult_col_letter}{r}"
        else:
            ws[f"{G}{r}"] = f"={F}{r}*1.3"
        ws[f"{H}{r}"] = f"={G}{r}/{E}{r}"
        ws[f"{CPCH}{r}"] = f"={K}{r}/{E}{r}"
        ws[f"{C40}{r}"] = f"={G}{r}*0.4"
        ws[f"{C50}{r}"] = f"={G}{r}*0.5"
        # L як формула (перезаписуємо значення, якщо були)
        ws[f"{L}{r}"] = f"={K}{r}/{G}{r}*100"

    # --- Conditional Formatting (оновлені правила) ---
    first_col_letter = get_column_letter(1)
    last_col_letter = get_column_letter(ws.max_column)
    data_range = f"{first_col_letter}{first_data_row}:{last_col_letter}{last_data_row}"

    grey = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Dynamic threshold per GEO (Габон=59, else 39)
    try:
        GEO_col_idx = headers.get("ГЕО") or headers.get("Geo") or headers.get("GEO") or headers.get("Країна")
        GEO = get_column_letter(GEO_col_idx) if GEO_col_idx else None
    except Exception:
        GEO = None
    THR = f'IF(${GEO}{first_data_row}="Габон",59,39)' if GEO else "39"

    # Grey: E = 0
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f"${E}{first_data_row}=0"], fill=grey, stopIfTrue=True),
    )

    # Green: INT(H) <= INT(I) AND L > 39
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[
                f"AND(${E}{first_data_row}>0,INT(${H}{first_data_row})<=INT(${I}{first_data_row}),${L}{first_data_row}>{THR})"],
            fill=green,
            stopIfTrue=True
        ),
    )

    # Yellow: (INT(H) <= INT(I)) OR (L > 39 AND H < I*1.31)
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[
                f"AND(${E}{first_data_row}>0,OR(INT(${H}{first_data_row})<=INT(${I}{first_data_row}),AND(${L}{first_data_row}>{THR},${H}{first_data_row}<${I}{first_data_row}*1.31)))"],
            fill=yellow,
            stopIfTrue=True
        ),
    )

    # Red: (E > 0 AND H > I*1.3 AND L > 39) OR (E > 0 AND INT(H) > INT(I) AND L < 39)
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[
                f"OR(AND(${E}{first_data_row}>0,${H}{first_data_row}>${I}{first_data_row}*1.3,${L}{first_data_row}>{THR}),AND(${E}{first_data_row}>0,INT(${H}{first_data_row})>INT(${I}{first_data_row}),${L}{first_data_row}<{THR}))"],
            fill=red,
            stopIfTrue=True
        ),
    )

    # --- Number format: 2 decimals for all numeric cells in data range
    # (краще з розділювачем тисяч)
    num_fmt = "#,##0.00"
    for row in ws.iter_rows(
            min_row=first_data_row,
            max_row=last_data_row,
            min_col=1,
            max_col=ws.max_column
    ):
        for cell in row:
            cell.number_format = num_fmt


def ask_additional_table_with_skip(message: types.Message, state: UserState):
    offer = state.offers[state.current_offer_index]
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton("Пропустити цей офер", callback_data="skip_offer"))
    bot.send_message(
        message.chat.id,
        (
            f"Надішліть додаткову таблицю для оферу:\n"
            f"<b>{offer}</b>\n\n"
            "Очікувані колонки: <b>Країна</b>, <b>Сума депозитів</b>.\n"
            "Або натисніть «Пропустити цей офер», щоб не включати його у фінальний звіт."
        ),
        reply_markup=kb,
    )


# ===================== ALLOCATION HELPERS =====================

def _classify_status(E: float, F: float, K: float) -> str:
    if E <= 0:
        return "Grey"
    H = 1.3 * F / E if E else float("inf")
    L = (100.0 * K) / (1.3 * F) if F else float("inf")

    # Green: E>0 and H<=9 and L>39.99
    green = (H <= H_THRESH + EPS) and (L > L_THRESH + EPS)
    if green:
        return "Green"

    # Yellow: E>0 and ((H<=9) or (L>39.99)) and not Green
    yellow = (H <= H_THRESH + EPS) or (L > L_THRESH + EPS)
    return "Yellow" if yellow else "Red"  # Red: H>9 and L<=39.99


def _fmt(v: float, suf: str = "", nan_text: str = "-") -> str:
    if not np.isfinite(v):
        return nan_text
    return f"{v:.2f}{suf}"


def build_allocation_explanation(df_source: pd.DataFrame,
                                 alloc_vec: pd.Series,
                                 budget: float,
                                 max_lines: int = 20) -> str:
    """
    Створює текстовий звіт:
      - скільки бюджету використано і залишок
      - скільки рядків змінили статус, скільки жовтих у підсумку
      - список топ-рядків з алокацією (Offer ID / Назва / ГЕО, +сума, нові H і L, статус: ДО → ПІСЛЯ)

    max_lines — обмеження на кількість детальних рядків у списку (щоб не перевантажувати чат).
    """
    df = df_source.copy()
    # нормалізуємо назви колонок на всяк випадок
    df.columns = [str(c).replace("\xa0", " ").replace("\u00A0", " ").strip() for c in df.columns]

    # Витягуємо потрібні колонки з дефолтами, якщо відсутні
    subid = df.get("Subid", pd.Series([""] * len(df)))
    offer = df.get("Offer ID", df.get("Назва Офферу", pd.Series([""] * len(df))))
    name = df.get("Назва Офферу", pd.Series([""] * len(df)))
    geo = df.get("ГЕО", pd.Series([""] * len(df)))
    E = pd.to_numeric(df.get("FTD qty", 0), errors="coerce").fillna(0.0)
    F = pd.to_numeric(df.get("Total spend", 0), errors="coerce").fillna(0.0)
    K = pd.to_numeric(df.get("Total Dep Amount", 0), errors="coerce").fillna(0.0)

    alloc = pd.to_numeric(alloc_vec, errors="coerce").reindex(df.index).fillna(0.0)
    F_new = (F + alloc)

    # Статуси ДО/ПІСЛЯ
    before = [_classify_status(float(E[i]), float(F[i]), float(K[i])) for i in df.index]
    after = [_classify_status(float(E[i]), float(F_new[i]), float(K[i])) for i in df.index]

    # Метрики
    total_budget = float(budget)
    used = float(alloc.sum())
    left = max(0.0, total_budget - used)

    yellow_before = sum(1 for s in before if s == "Yellow")
    yellow_after = sum(1 for s in after if s == "Yellow")
    green_to_yellow = sum(1 for i in df.index if (before[i] == "Green" and after[i] == "Yellow"))

    # Побудова списку рядків з алокацією
    rows = []
    for i in alloc.index:
        if alloc[i] <= 0:
            continue

        Ei = float(E[i]);
        Fi = float(F[i]);
        Ki = float(K[i]);
        Fni = float(F_new[i])

        # ДО
        H_before = (1.3 * Fi / Ei) if Ei > 0 else float("inf")
        L_before = (100.0 * Ki) / (1.3 * Fi) if Fi > 0 else float("inf")

        # ПІСЛЯ
        H_after = (1.3 * Fni / Ei) if Ei > 0 else float("inf")
        L_after = (100.0 * Ki) / (1.3 * Fni) if Fni > 0 else float("inf")

        line = (
            f"- {str(offer[i]) or ''} / {str(name[i]) or ''} / {str(geo[i]) or ''}: "
            f"+{alloc[i]:.2f} → Total Spend {Fi:.2f}→{Fni:.2f}; "
            f"CPA {_fmt(H_before)}→{_fmt(H_after)}, "
            f"My deposit amount {_fmt(L_before, '%')}→{_fmt(L_after, '%')} | "
            f"{before[i]} → {after[i]}"
        )
        rows.append((alloc[i], line))

    # Сортуємо за найбільшою алокацією і обрізаємо
    rows.sort(key=lambda x: (-x[0], x[1]))
    detail_lines = [ln for _, ln in rows[:max_lines]]

    header = (
        f"Розподіл бюджету: {used:.2f} / {total_budget:.2f} використано; залишок {left:.2f}\n"
        f"Жовтих ДО/ПІСЛЯ: {yellow_before} → {yellow_after} (зел.→жовт.: {green_to_yellow})"
    )

    if not detail_lines:
        return header + "\n\n(Алокації по рядках відсутні — бюджет не було куди розподілити за правилами.)"

    return header + "\n\nТоп розподілів:\n" + "\n".join(detail_lines) + \
        ("\n\n…Список обрізано." if len(rows) > max_lines else "")


def allocate_with_openai(df: pd.DataFrame, rules_text: str, model: str | None = None) -> pd.DataFrame:
    """
    Надсилає таблицю і правила в OpenAI, отримує назад оновлену таблицю.
    Модель має ПОВЕРНУТИ CSV з тими ж колонками + колонка NEW SPEND (або оновити target-колонку).
    """
    if df.empty:
        raise ValueError("Пуста таблиця для алокації.")

    model = model or OPENAI_MODEL
    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

    # 1) Ділимо великий DF на чанки
    chunks = _split_df_by_size(df)

    updated_chunks: list[pd.DataFrame] = []

    for idx, chunk in enumerate(chunks, start=1):
        csv_in = _df_to_csv(chunk)

        system_msg = (
            "You are a meticulous data allocator. "
            "Follow the given allocation rules exactly. "
            "Return ONLY a CSV table with the SAME columns as input, "
            f"and include a numeric column '{OPENAI_OUTPUT_COLUMN}' with the recomputed spend per row. "
            "Do not add extra commentary. Use dot as decimal separator."
        )

        user_msg = (
            "Rules:\n"
            f"{rules_text}\n\n"
            "Instructions:\n"
            f"- Input table is a CSV.\n"
            f"- Keep all original columns unchanged.\n"
            f"- Add or overwrite a column named '{OPENAI_OUTPUT_COLUMN}' with the new per-row allocation (number only).\n"
            "- Return ONLY the CSV (no explanations).\n\n"
            "Input CSV:\n"
            f"{csv_in}"
        )

        resp = client.chat.completions.create(
            model=model,
            temperature=0.0,
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user", "content": user_msg},
            ],
        )

        content = resp.choices[0].message.content or ""
        csv_out = _csv_from_text(content)

        # читаємо назад у DF
        try:
            out_df = pd.read_csv(io.StringIO(csv_out))
        except Exception as e:
            raise RuntimeError(f"Не вдалося розпарсити CSV від моделі на чанку {idx}: {e}")

        # базова валідація
        missing_cols = [c for c in chunk.columns if c not in out_df.columns]
        if missing_cols:
            raise RuntimeError(f"Модель не повернула всі колонки (чанк {idx}). Відсутні: {missing_cols}")

        if OPENAI_OUTPUT_COLUMN not in out_df.columns:
            raise RuntimeError(f"Модель не повернула колонку '{OPENAI_OUTPUT_COLUMN}' (чанк {idx}).")

        # Зберігаємо порядок рядків: приєднаємо по індексу
        # (очікується той самий порядок — але на всякий випадок приведемо довжини)
        if len(out_df) != len(chunk):
            # як fallback — підрізаємо/доповнювати не будемо; вважаємо помилкою
            raise RuntimeError(
                f"Розмір чанка змінився (очікував {len(chunk)}, отримав {len(out_df)}) на чанку {idx}."
            )

        # беремо тільки колонку з новими значеннями і змерджимо
        chunk[OPENAI_OUTPUT_COLUMN] = out_df[OPENAI_OUTPUT_COLUMN].values
        updated_chunks.append(chunk)

    # збираємо назад
    result = pd.concat(updated_chunks, axis=0)
    result.reset_index(drop=True, inplace=True)
    return result


def allocate_total_spend_alternative(
        df: pd.DataFrame,
        *,
        col_total_spend: str = "Total spend",  # F
        col_ftd_qty: str = "FTD qty",  # E
        col_cpa_target: str = "CPA Target",  # I
        col_my_deposit: str = "My deposit amount",  # L (we reuse as a scratch to show computed L_now)
        col_total_dep_amount: str = "Total Dep Amount",  # K
        col_geo: str = "ГЕО",
        col_offer: str = "Назва Офферу",
        plus35_pairs: Optional[set[tuple[str, str]]] = None,  # set of (_norm_pair(offer, geo))
        in_place: bool = False,
        round_decimals: Optional[int] = 2,
        excel_path: Optional[str] = None,
        sheet_name: str = "Result",
        header_row: int = 1,
        skip_pass2: bool = False,  # if True: stop after Pass#1 + L-adjust; return leftover
        return_leftover: bool = False,  # if True: return (df, leftover_budget)
) -> Union[pd.DataFrame, tuple[pd.DataFrame, float]]:
    """
    Alternative allocation:
      1) budget = sum(F) over all rows
      2) zero all F
      3) allocate only over rows with Current='+' and E>0 (FTD>0)
    Pass#1 -> 'yellow' target
    Adjust L>THR with two caps (per-row converter depends on plus35: 100/135 or 100/130)
    Pass#2 -> 'red' ceiling (skipped if skip_pass2=True)

    If return_leftover=True, returns (df, budget_left) instead of df.
    """

    work = df if in_place else df.copy()

    # --- sanity ---
    for c in (col_total_spend, col_ftd_qty, col_cpa_target, col_my_deposit, col_total_dep_amount):
        if c not in work.columns:
            raise KeyError(f"Відсутня колонка: {c}")
    if col_geo not in work.columns:
        raise KeyError(f"Відсутня колонка з ГЕО: {col_geo}")
    if col_offer not in work.columns:
        for alt in ("Назва Офферу", "Оффер", "Offer", "Назва оффера", "Назва офферу"):
            if alt in work.columns:
                col_offer = alt
                break

    # --- types ---
    for c in (col_total_spend, col_ftd_qty, col_cpa_target, col_my_deposit, col_total_dep_amount):
        work[c] = pd.to_numeric(work[c], errors="coerce").fillna(0.0)

    # --- budget & zeroing ---
    budget = float(work[col_total_spend].sum())
    if budget < 0:
        budget = 0.0
    work[col_total_spend] = 0.0

    # --- Current column & scope ---
    # We require a valid Current/current column so the scope is intentional.
    cur_col = None
    for cand in ("Current", "current"):
        if cand in work.columns:
            cur_col = cand
            break
    if cur_col is None:
        # If you prefer a fallback instead of error, replace the next line with:
        # cur_col = "Current"; work[cur_col] = "+"
        raise ValueError("Не знайдено колонку 'Current'/'current'. Спочатку додайте її (з '+' для потрібних рядків).")

    mask_current = (work[cur_col].astype(str).str.strip() == "+")
    mask_ftd = (work[col_ftd_qty] > 0)
    mask_scope = (mask_current & mask_ftd)

    if not bool(mask_scope.any()):
        # If you prefer to fallback silently to all E>0 rows, replace 'raise' with:
        # mask_scope = mask_ftd
        # (and maybe log a warning). For now, we error to avoid silent all-zero allocations.
        raise ValueError(
            "Немає рядків для алокації: потрібні Current='+' та FTD qty > 0. "
            "Перевірте колонку Current та значення FTD."
        )

    # --- constants ---
    CONV = 100.0 / 130.0  # used in Pass#1 and red ceiling (per your spec)
    MULT_Y_HI = 1.3
    MULT_Y_LO = 1.1
    MULT_CPA_Y = 1.3
    MULT_RED = 1.8

    # threshold per GEO
    geo_raw = work[col_geo].astype(str).str.strip().fillna("")
    L_threshold_series = np.where(geo_raw.eq("Габон"), 59.0, 39.0)

    # Use original L (My deposit amount) to choose yellow multiplier in Pass#1
    L_for_threshold = pd.to_numeric(df[col_my_deposit], errors="coerce").fillna(0.0).clip(lower=0)

    # ---------------- Pass#1: 'yellow' ----------------
    idx_pass1 = work.loc[mask_scope].sort_values(by=col_ftd_qty, ascending=True).index
    for i in idx_pass1:
        if budget <= 0:
            break
        E = float(work.at[i, col_ftd_qty])
        I = float(work.at[i, col_cpa_target])
        Lthr_val = float(L_for_threshold.at[i])
        thr_i = float(L_threshold_series[work.index.get_loc(i)])

        mult = MULT_Y_HI if Lthr_val >= thr_i else MULT_Y_LO
        target_F = E * I * mult * CONV
        add = min(target_F, budget)
        work.at[i, col_total_spend] = add
        budget -= add

    # -------- Recompute L, adjust where L>THR with 2 caps --------
    # L_now = (K / (F*1.3)) * 100
    G_now = work[col_total_spend] * 1.3
    with np.errstate(divide='ignore', invalid='ignore'):
        L_now = np.where(G_now > 0, (work[col_total_dep_amount] / G_now) * 100.0, np.inf)
    work[col_my_deposit] = L_now  # show in L column for visibility

    mask_adjust = mask_scope.values & (L_now > L_threshold_series)

    # === per-row CONV for caps if plus35 ===
    if plus35_pairs:
        def _n(s: str) -> str:
            s = str(s or "").strip().lower()
            s = s.replace("’", "'").replace("`", "'").replace("–", "-").replace("—", "-")
            s = re.sub(r"\s+", " ", s)
            return s

        offer_norm = work[col_offer].map(_n) if col_offer in work.columns else pd.Series([""] * len(work),
                                                                                         index=work.index)
        geo_norm = work[col_geo].map(_n)
        is_plus35 = np.array([(o, g) in plus35_pairs for o, g in zip(offer_norm, geo_norm)], dtype=bool)
    else:
        is_plus35 = np.zeros(len(work), dtype=bool)

    # per-row converters for the caps ONLY (as requested)
    CONV_CAP = np.where(is_plus35, 100.0 / 135.0, 100.0 / 130.0)

    if mask_adjust.any():
        K = work[col_total_dep_amount].values
        E = work[col_ftd_qty].values
        I = work[col_cpa_target].values
        TH = L_threshold_series

        # caps honor plus35 via CONV_CAP
        F_cap_dep = (K / TH * 100.0) * CONV_CAP
        F_cap_cpa = (E * I * MULT_CPA_Y) * CONV_CAP
        F_target = np.minimum(F_cap_dep, F_cap_cpa)

        curF = work[col_total_spend].values
        adj = mask_adjust

        # raise where needed (consume budget)
        need_up_mask = adj & (F_target > curF)
        delta_up = F_target - curF
        need_up_total = float(delta_up[need_up_mask].sum())
        if need_up_total > 0 and budget > 0:
            ratio = min(1.0, budget / need_up_total)
            inc = np.zeros_like(curF, dtype=float)
            inc[need_up_mask] = delta_up[need_up_mask] * ratio
            curF += inc
            budget -= float(inc.sum())

        # lower where needed (free budget)
        need_down_mask = adj & (F_target < curF)
        freed = float((curF[need_down_mask] - F_target[need_down_mask]).sum())
        if freed > 0:
            curF[need_down_mask] = F_target[need_down_mask]
            budget += freed

        work[col_total_spend] = curF

        # refresh L display
        G_now = work[col_total_spend] * 1.3
        with np.errstate(divide='ignore', invalid='ignore'):
            L_now = np.where(G_now > 0, (work[col_total_dep_amount] / G_now) * 100.0, np.inf)
        work[col_my_deposit] = L_now

    # ---------------- Pass#2: 'red' ceiling ----------------
    if (not skip_pass2) and budget > 0:
        idx_pass2 = work.loc[mask_scope].sort_values(by=col_total_spend, ascending=True).index
        for i in idx_pass2:
            if budget <= 0:
                break
            E = float(work.at[i, col_ftd_qty])
            I = float(work.at[i, col_cpa_target])
            red_cap = E * I * MULT_RED * CONV
            curF = float(work.at[i, col_total_spend])
            need = max(0.0, red_cap - curF)
            if need <= 0:
                continue
            add = min(need, budget)
            work.at[i, col_total_spend] = curF + add
            budget -= add

    # rounding
    if round_decimals is not None:
        work[col_total_spend] = work[col_total_spend].round(round_decimals)

    # write excel if requested
    if excel_path:
        with ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
            work.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.book[sheet_name]
            first_data_row = header_row + 1
            last_data_row = header_row + len(work)
            inject_formulas_and_cf(ws, header_row=header_row, first_data_row=first_data_row,
                                   last_data_row=last_data_row)
            # (optional) here you can re-write 'Total+%' per row to 1.35 for plus35 pairs if you need
            writer._save()

    return (work, budget) if return_leftover else work


def compute_optimal_allocation(df: pd.DataFrame, budget: float) -> Tuple[pd.DataFrame, str, pd.Series]:
    """
    Нова послідовність:
      A) Спочатку намагаємось мінімально перевести GREEN -> YELLOW (дотримуючись CPA<=CPA_CAP).
      B) Якщо залишився бюджет — насичуємо YELLOW максимально, але так, щоб вони залишались YELLOW (і CPA<=CPA_CAP).

    Позначення:
      E = FTD qty
      F = Total spend
      K = Total Dep Amount

    Межі/похідні:
      F_at_H = H_THRESH * E / 1.3
      F_at_L = (100 * K) / (1.3 * L_THRESH)  # L == L_THRESH при такому F
      F_cap  = CPA_CAP * E / 1.3
    """
    dfw = df.copy()

    # Числові колонки
    E = pd.to_numeric(dfw["FTD qty"], errors="coerce").fillna(0.0)
    F = pd.to_numeric(dfw["Total spend"], errors="coerce").fillna(0.0)
    K = pd.to_numeric(dfw["Total Dep Amount"], errors="coerce").fillna(0.0)

    # Поточні H, L
    with np.errstate(divide='ignore', invalid='ignore'):
        H = 1.3 * F / E.replace(0, np.nan)
        L = 100.0 * K / (1.3 * F.replace(0, np.nan))

    # Межі
    F_at_H = H_THRESH * E / 1.3
    F_at_L = (100.0 * K) / (1.3 * L_THRESH)
    F_cap = CPA_CAP * E / 1.3

    # Маски статусів (строго відповідно до правил/Excel)
    grey_mask = (E <= 0)
    green_mask = (~grey_mask) & (H <= H_THRESH + EPS) & (L > L_THRESH + EPS)
    yellow_mask = (~grey_mask) & ((H <= H_THRESH + EPS) | (L > L_THRESH + EPS)) & (~green_mask)
    # red_mask   = (~grey_mask) & (~green_mask) & (~yellow_mask)  # не потрібен явно

    alloc = pd.Series(0.0, index=dfw.index, dtype=float)
    rem = float(budget) if budget and budget > 0 else 0.0

    # -------------------------------
    # A) GREEN -> YELLOW (мінімальний spend)
    # -------------------------------
    # Кандидати цільових F:
    #   - перетнути межу H: F_cross_H = F_at_H + EPS_YEL (робить H трохи > H_THRESH)
    #   - перетнути межу L: F_cross_L = F_at_L + EPS_YEL (робить L трохи < L_THRESH)
    F_cross_H = F_at_H + EPS_YEL
    F_cross_L = F_at_L + EPS_YEL

    # Мінімальний F, який зламав "зеленість", але не робить рядок "червоним" і не перевищує CPA cap.
    candidates = pd.DataFrame({
        "F_now": F,
        "F_cap": F_cap,
        "F_cross_H": F_cross_H,
        "F_cross_L": F_cross_L,
        "E": E,
        "K": K
    })

    # Для кожного green обчислюємо найменшу допустиму ціль F_target
    F_target = F.copy()

    for i in candidates[green_mask].index:
        Fi = float(candidates.at[i, "F_now"])
        Fcap = float(candidates.at[i, "F_cap"])
        Fh = float(candidates.at[i, "F_cross_H"])
        Fl = float(candidates.at[i, "F_cross_L"])
        Ei = float(E.at[i])
        Ki = float(K.at[i])

        # Обидва потенційні цілі в межах CPA?
        options = []
        for Ft in (Fh, Fl):
            if np.isfinite(Ft) and Ft > Fi + EPS and Ft <= Fcap + EPS:
                # Перевіримо, що Ft не робить рядок "червоним"
                Ht = 1.3 * Ft / Ei if Ei > 0 else float("inf")
                Lt = (100.0 * Ki) / (1.3 * Ft) if Ft > 0 else float("inf")
                is_red = (Ht > H_THRESH + EPS) and (Lt <= L_THRESH + EPS)
                if not is_red:
                    options.append(Ft)

        if options:
            F_target.at[i] = min(options)  # найменша ціна переходу
        else:
            # неможливо легально зробити жовтим — залишаємо як є
            F_target.at[i] = Fi

    need_delta = (F_target - F).clip(lower=0.0)

    # Розподіл: за зростанням потрібної дельти
    for i in need_delta[green_mask].sort_values(ascending=True).index:
        if rem <= 1e-9:
            break
        need = float(need_delta.at[i])
        if need <= 0:
            continue
        take = min(rem, need)
        alloc.at[i] += take
        rem -= take

    # -------------------------------
    # B) Насичення YELLOW, щоб лишались YELLOW (CPA<=cap)
    # -------------------------------
    if rem > 1e-9:
        # Перерахувати F після кроку A
        F_mid = F + alloc
        with np.errstate(divide='ignore', invalid='ignore'):
            H_mid = 1.3 * F_mid / E.replace(0, np.nan)
            L_mid = 100.0 * K / (1.3 * F_mid.replace(0, np.nan))

        # Ті, хто зараз жовті (включно з новими з кроку A)
        is_green_mid = (~(E <= 0)) & (H_mid <= H_THRESH + EPS) & (L_mid > L_THRESH + EPS)
        is_yellow_mid = (~(E <= 0)) & (((H_mid <= H_THRESH + EPS) | (L_mid > L_THRESH + EPS)) & (~is_green_mid))

        # Межа "залишитись жовтим": до max(F_at_H, F_at_L - EPS_YEL), та ще й не перевищити cap
        F_yellow_limit_base = pd.Series(np.maximum(F_at_H, F_at_L - EPS_YEL), index=dfw.index)
        F_yellow_limit_final = pd.Series(np.minimum(F_yellow_limit_base, F_cap), index=dfw.index).fillna(0.0)

        headroom = (F_yellow_limit_final - F_mid).clip(lower=0.0)

        # Greedy за спаданням headroom
        for i in headroom[is_yellow_mid].sort_values(ascending=False).index:
            if rem <= 1e-9:
                break
            give = float(min(rem, headroom.at[i]))
            if give <= 0:
                continue
            alloc.at[i] += give
            rem -= give

    # ПІДСУМОК
    F_final = F + alloc
    with np.errstate(divide='ignore', invalid='ignore'):
        H_final = 1.3 * F_final / E.replace(0, np.nan)
        L_final = 100.0 * K / (1.3 * F_final.replace(0, np.nan))

    still_green = (E > 0) & (H_final <= H_THRESH + EPS) & (L_final > L_THRESH + EPS)
    still_yellow = (E > 0) & (((H_final <= H_THRESH + EPS) | (L_final > L_THRESH + EPS)) & (~still_green))

    kept_yellow = int(still_yellow.sum())
    total_posE = int((E > 0).sum())

    dfw["Allocated extra"] = alloc
    dfw["New Total spend"] = F_final
    dfw["Will be yellow"] = ["Yes" if x else "No" for x in still_yellow]

    summary = (
        f"Бюджет: {budget:.2f}\n"
        f"Жовтих після розподілу: {kept_yellow}/{total_posE}\n"
        f"Правила: спочатку переводимо зелені в жовті мінімальним spend (CPA≤{CPA_CAP:g}), "
        f"потім насичуємо жовті в межах жовтого (H≤{H_THRESH:g} або L>{L_THRESH:.2f}, CPA≤{CPA_CAP:g})."
    )
    return dfw, summary, alloc


def write_result_like_excel_with_new_spend(bio: io.BytesIO, df_source: pd.DataFrame, new_total_spend: pd.Series):
    """
    Build an Excel sheet identical to result.xlsx structure:
    Columns (A..P):
      A Subid | B Offer ID | C Назва Офферу | D ГЕО | E FTD qty | F Total spend | G Total+% | H CPA | I CPA Target |
      J СP/Ч | K Total Dep Amount | L My deposit amount | M C. profit Target 40% | N C. profit Target 50% | O CAP | P Остаток CAP
    Uses new_total_spend for column F, then writes the same formulas, formats and conditional formatting.
    """
    final_cols = [
        "Subid", "Offer ID", "Назва Офферу", "ГЕО",
        "FTD qty", "Total spend", "Total+%", "CPA", "CPA Target", "СP/Ч",
        "Total Dep Amount", "My deposit amount", "C. profit Target 40%", "C. profit Target 50%",
        "CAP", "Остаток CAP", "Current"
    ]

    # Start from source, ensure all columns exist
    df_out = df_source.copy()
    # normalize column names just in case
    df_out.columns = [str(c).replace("\xa0", " ").replace("\u00A0", " ").strip() for c in df_out.columns]

    # Coerce numbers
    df_out["FTD qty"] = pd.to_numeric(df_out.get("FTD qty", 0), errors="coerce").fillna(0)
    df_out["Total spend"] = pd.to_numeric(df_out.get("Total spend", 0), errors="coerce").fillna(0.0)
    df_out["Total Dep Amount"] = pd.to_numeric(df_out.get("Total Dep Amount", 0.0), errors="coerce").fillna(0.0)

    # Apply new spend
    # align by index; if shapes don't match, reindex new_total_spend to df_out
    new_total_spend = pd.to_numeric(new_total_spend, errors="coerce").fillna(0.0)
    new_total_spend = new_total_spend.reindex(df_out.index).fillna(0.0)
    df_out["Total spend"] = (df_out["Total spend"] + new_total_spend).round(2)

    # Ensure missing columns exist as blanks
    for col in final_cols:
        if col not in df_out.columns:
            df_out[col] = ""

    # Reorder
    df_out = df_out[final_cols].copy()

    # Round numeric display columns
    df_out["FTD qty"] = pd.to_numeric(df_out["FTD qty"], errors="coerce").fillna(0).astype(int)
    for col in ["Total spend", "Total Dep Amount"]:
        df_out[col] = pd.to_numeric(df_out[col], errors="coerce").round(2)

    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Result")
        wb = writer.book
        ws = writer.sheets["Result"]

        from openpyxl.styles import PatternFill, Alignment, Font
        from openpyxl.formatting.rule import FormulaRule

        first_row = 2
        last_row = ws.max_row

        # Re-insert formulas (same as send_final_table)
        for r in range(first_row, last_row + 1):
            offer_val = ws[f"C{r}"].value
            geo_val = ws[f"D{r}"].value
            mul = 1.3
            if plus35_pairs:
                try:
                    if _norm_pair(offer_val, geo_val) in plus35_pairs:
                        mul = 1.35
                except Exception:
                    pass
            ws[f"G{r}"].value = f"=F{r}*{mul}"  # Total+%
            ws[f"H{r}"].value = f"=IFERROR(G{r}/E{r},\"\")"  # CPA
            ws[f"I{r}"].value = cpa_target_for_geo(ws[f"D{r}"].value)  # CPA Target
            ws[f"J{r}"].value = f"=IFERROR(K{r}/E{r},\"\")"  # СP/Ч
            ws[f"L{r}"].value = f"=IFERROR(K{r}/G{r}*100,0)"  # My deposit amount
            ws[f"M{r}"].value = f"=G{r}*0.4"  # Profit 40%
            ws[f"N{r}"].value = f"=G{r}*0.5"  # Profit 50%

        # Header styling
        for col in range(1, 17):  # A..P
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

        # Widths
        widths = {"A": 10, "B": 12, "C": 22, "D": 16, "E": 10, "F": 14, "G": 12, "H": 10, "I": 12, "J": 10, "K": 16,
                  "L": 18, "M": 18, "N": 18, "O": 12, "P": 16}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # Number formats (2 decimals for F,G,H,J,K,L,M,N; E as integer)
        for r in range(first_row, last_row + 1):
            ws[f"E{r}"].number_format = "0"
        two_dec_cols = ["F", "G", "H", "J", "K", "L", "M", "N"]
        for col in two_dec_cols:
            for r in range(first_row, last_row + 1):
                ws[f"{col}{r}"].number_format = "0.00"

        # Conditional formatting — SAME rules/colors
        data_range = f"A{first_row}:P{last_row}"
        grey = PatternFill("solid", fgColor="BFBFBF")
        green = PatternFill("solid", fgColor="C6EFCE")
        yellow = PatternFill("solid", fgColor="FFEB9C")
        red = PatternFill("solid", fgColor="FFC7CE")

        THR2 = 'IF($D2="Габон",59,39)'

        # Grey: E = 0
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=["$E2=0"], fill=grey, stopIfTrue=True)
        )

        # Green: INT(H) <= INT(I) AND L > 39
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f"AND($E2>0,INT($H2)<=INT($I2),$L2>{THR2})"], fill=green, stopIfTrue=True)
        )

        # Yellow: (INT(H) <= INT(I)) OR (L > 39 AND H < I*1.31)
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f"AND($E2>0,OR(INT($H2)<=INT($I2),AND($L2>{THR2},$H2<$I2*1.31)))"], fill=yellow,
                        stopIfTrue=True)
        )

        # Red: (E > 0 AND H > I*1.3 AND L > 39) OR (E > 0 AND INT(H) > INT(I) AND L < 39)
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f"OR(AND($E2>0,$H2>$I2*1.3,$L2>{THR2}),AND($E2>0,INT($H2)>INT($I2),$L2<{THR2}))"],
                fill=red,
                stopIfTrue=True
            )
        )


def _run_alternative_and_send(chat_id: int, state: UserState):
    """
    Runs allocate_total_spend_alternative with optional pairs
    and sends 'allocation_alternative.xlsx' back to the user.
    Resets phase to WAIT_MAIN at the end.
    """
    if state.alloc_df is None:
        bot.send_message(chat_id, "Немає завантаженої таблиці Result. Спробуйте /allocate ще раз.")
        state.phase = "WAIT_MAIN"
        return

    out_df, leftover_budget = allocate_total_spend_alternative(
        state.alloc_df,
        col_total_spend="Total spend",
        col_ftd_qty="FTD qty",
        col_cpa_target="CPA Target",
        col_my_deposit="My deposit amount",
        col_total_dep_amount="Total Dep Amount",
        col_geo="ГЕО",
        col_offer="Назва Офферу",
        plus35_pairs=getattr(state, 'plus35_pairs', None),
        excel_path="Result.xlsx",
        sheet_name="Result",
        header_row=1,
        return_leftover=True,
    )

    try:
        with open("Result.xlsx", "rb") as f:
            bot.send_document(
                chat_id,
                f,
                visible_file_name="allocation_alternative.xlsx",
                caption=(
                    "Готово: алокація за альтернативним режимом"
                )
            )
            bot.send_message(chat_id, f'Залишок бюджету після розподілу: $<b>{leftover_budget:,.2f}</b>',
                             parse_mode="HTML")
    except Exception as e:
        bot.send_message(chat_id, f"⚠️ Не вдалося відправити Excel: <code>{e}</code>", parse_mode="HTML")
    finally:
        # reset minimal allocate state
        state.phase = "WAIT_MAIN"
        state.alloc_df = None


# ===================== PAIR HELPERS =====================
def _norm_pair(offer: str, geo: str) -> tuple[str, str]:
    """
    Нормалізація пари (Назва Офферу, ГЕО):
    - lowercase
    - strip spaces
    - заміна різних апострофів, дефісів, нестандартних лапок
    - нормалізація unicode (ʼ → ')
    """

    def clean(s: str) -> str:
        if not s:
            return ""
        s = str(s).strip().lower()
        s = unicodedata.normalize("NFKC", s)
        s = s.replace("’", "'").replace("ʼ", "'").replace("`", "'")
        s = re.sub(r"[\u2013\u2014]", "-", s)  # em/en dash → -
        s = re.sub(r"\s+", " ", s)
        return s

    return (clean(offer), clean(geo))


def _mark_current(offer: str, geo: str, current_pairs: set[tuple[str, str]] | None) -> str:
    if not current_pairs:
        return "-"  # якщо current.xlsx не завантажено — дефолтно '-'
    return "+" if _norm_pair(offer, geo) in current_pairs else "-"


def _extract_pairs_df(df: pd.DataFrame) -> set[tuple[str, str]]:
    lower = {str(c).strip().lower(): c for c in df.columns}
    geo_col = lower.get('гео') or lower.get('geo')
    offer_col = lower.get('назва офферу') or lower.get('назва оферу') or lower.get('offer') or lower.get('offer id')
    if not geo_col or not offer_col:
        raise ValueError("Очікую дві колонки: 'Назва Офферу' і 'ГЕО' (або 'Гео').")
    pairs = set()
    for _, row in df[[offer_col, geo_col]].dropna().iterrows():
        pairs.add(_norm_pair(row[offer_col], row[geo_col]))
    return pairs


def build_en_to_uk_preferred(uk_to_en: Dict[str, str]) -> Dict[str, str]:
    """
    Побудувати EN -> UA мапу з пріоритетом українських назв із вашої UA-мапи.
    Ключі у normalize_text, значення — як є (красиві UA назви).
    """
    en_to_uk: Dict[str, str] = {}
    for ua_raw, en_raw in uk_to_en.items():
        ua_key = normalize_text(ua_raw)
        en_key = normalize_text(en_raw)
        # Перший, хто запише, переможе (стабільна перевага)
        en_to_uk.setdefault(en_key, ua_raw)
    return en_to_uk


def unite_ua_ru_rows(
        df: pd.DataFrame,
        *,
        country_col: str = "ГЕО",
        total_spend_col: str = "Total spend",
        offer_col: str = "Назва Офферу",
) -> pd.DataFrame:
    """
    Об'єднує рядки UA/RU країни ТІЛЬКИ в межах одного оффера:
      ключ = (offer_col, canonical_en(country))
    Сумуємо ЛИШЕ Total spend (все інше — як у першому рядку групи).
    Повертаємо повну таблицю з усіма наявними колонками у вхідному df.
    """

    # Мапи
    uk_to_en = build_country_map_uk_to_en()
    canonical = build_country_canonical()
    try:
        ru_to_en = build_country_map_ru_to_en()
    except NameError:
        ru_to_en = None

    en_to_uk = build_en_to_uk_preferred(uk_to_en)

    work = df.copy()

    # Підхопити альтернативні назви сумової колонки (BUDG кейс)
    if total_spend_col not in work.columns and "Загальні витрати" in work.columns:
        total_spend_col = "Загальні витрати"

    if country_col not in work.columns:
        raise KeyError(f"Відсутня колонка країни: {country_col}")
    if offer_col not in work.columns:
        raise KeyError(f"Відсутня колонка оффера: {offer_col}")

    # Канон EN країни (UA->EN, RU->EN, canonical)
    work["__canon_en__"] = work[country_col].apply(
        lambda x: to_canonical_en(x, uk_to_en, canonical, ru_to_en)
    )

    # Норм-ключ для групування
    def _nz(s):
        return normalize_text(s)

    work["__canon_key__"] = work["__canon_en__"].map(_nz)
    work["__offer_key__"] = work[offer_col].astype(str).map(_nz)

    # Аггрегація: Total spend -> sum; решта -> first
    agg_dict = {total_spend_col: "sum"}
    # зберігаємо всі інші колонки (включно з формульними, значення візьмемо з першого рядка)
    for c in work.columns:
        if c not in (total_spend_col, "__canon_en__", "__canon_key__", "__offer_key__"):
            agg_dict.setdefault(c, "first")

    out = (
        work.groupby(["__offer_key__", "__canon_key__"], as_index=False)
        .agg(agg_dict)
    )

    # Відновлюємо українську назву країни: пріоритет з UA-мапи, інакше беремо першу з групи
    def _uk_name(row):
        ck = row["__canon_key__"]
        if ck in en_to_uk:
            return en_to_uk[ck]
        # знайти оригінальну з групи
        mask = (work["__canon_key__"] == ck) & (work["__offer_key__"] == row["__offer_key__"])
        sample = work.loc[mask, country_col].dropna().astype(str)
        return sample.iloc[0] if not sample.empty else row.get("__canon_en__", "")

    out[country_col] = out.apply(_uk_name, axis=1)

    # Прибрати службові
    for c in ("__canon_en__", "__canon_key__", "__offer_key__"):
        if c in out.columns:
            out.drop(columns=[c], inplace=True)

    # Привести число
    out[total_spend_col] = pd.to_numeric(out[total_spend_col], errors="coerce").fillna(0).round(2)

    return out


# ===================== BOT HANDLERS =====================

@bot.message_handler(commands=["start", "help"])
@require_access
def cmd_start(message: types.Message):
    chat_id = message.chat.id

    # ✅ ЗБЕРЕЖЕННЯ ДАНИХ КОРИСТУВАЧА (id, first_name, last_name, username)
    save_user_if_new(message.from_user)

    st = user_states.setdefault(chat_id, UserState())

    if getattr(st, "current_pairs", None) is None:
        st.current_pairs = GLOBAL_CURRENT_PAIRS
    if getattr(st, "plus35_pairs", None) is None:
        st.plus35_pairs = GLOBAL_PLUS35_PAIRS

    has_current = bool(st.current_pairs)

    # 🧭 Скидаємо робочі фази/буфери
    st.alloc_mode = None
    st.main_agg_df = None
    st.offers = []
    st.current_offer_index = 0

    if has_current:
        # ✅ current уже відомий — одразу просимо головну таблицю
        st.phase = "WAIT_MAIN"
        bot.reply_to(
            message,
            (
                "Привіт! 👋\n\n"
                "Можна одразу надсилати головну таблицю (CSV) — аркуш BUDG з колонками: Назва Офферу, ГЕО, Загальні витрати."
            ),
            parse_mode="HTML",
        )
    else:
        # ❗ current ще не завантажений — просимо вперше
        st.phase = "WAIT_CURRENT"
        bot.reply_to(
            message,
            (
                "Привіт! 👋\n\n"
                "Спочатку надішліть файл <b>current.xlsx</b>/<b>.csv</b> з колонками "
                "<i>Назва Офферу</i> і <i>ГЕО</i> — я збережу пари та додаватиму колонку <b>Current</b> (+/−). "
                "Це одноразове налаштування: надалі повторно надсилати файл не потрібно.\n\n"
                "Після цього — надішліть головну таблицю BUDG."
            ),
            parse_mode="HTML",
        )


@bot.message_handler(content_types=["document"])
@require_access
def on_document(message: types.Message):
    import pandas as pd

    chat_id = message.chat.id
    state = user_states.setdefault(chat_id, UserState())

    try:
        # === 1. Обробка файлів current.xlsx / plus35.xlsx ===
        if state.phase in ("WAIT_CURRENT", "WAIT_PLUS35"):
            filename = message.document.file_name or "file"
            file_info = bot.get_file(message.document.file_id)
            file_bytes = bot.download_file(file_info.file_path)

            # зчитування XLSX / CSV
            import io, pandas as pd
            if filename.lower().endswith((".xlsx", ".xls")):
                dfp = pd.read_excel(io.BytesIO(file_bytes))
            elif filename.lower().endswith(".csv"):
                dfp = pd.read_csv(io.BytesIO(file_bytes))
            else:
                bot.reply_to(message, "⚠️ Підтримуються лише .xlsx/.xls/.csv файли")
                return

            # уніфіковане витягнення пар
            lower = {str(c).strip().lower(): c for c in dfp.columns}
            offer_col = lower.get("назва офферу") or lower.get("offer") or lower.get("оффер")
            geo_col = lower.get("гео") or lower.get("geo") or lower.get("країна")

            if not (offer_col and geo_col):
                bot.reply_to(message, "❌ У файлі мають бути колонки 'Назва Офферу' і 'ГЕО'")
                return

            def _norm(s: str) -> str:
                s = str(s or "").strip().lower()
                s = s.replace("’", "'").replace("`", "'").replace("–", "-").replace("—", "-")
                s = re.sub(r"\s+", " ", s)
                return s

            pairs = {
                (_norm(r[offer_col]), _norm(r[geo_col]))
                for _, r in dfp.dropna(subset=[offer_col, geo_col]).iterrows()
            }

            # збереження в state
            if state.phase == "WAIT_CURRENT":
                state.current_pairs = pairs
                # update globals + persist
                globals()["GLOBAL_CURRENT_PAIRS"] = pairs
                save_pairs()

                bot.reply_to(
                    message,
                    f"✅ Збережено {len(pairs)} поточних пар. Можна надсилати головну таблицю (BUDG).",
                    parse_mode="HTML",
                )
                state.phase = "WAIT_MAIN"
                return

            elif state.phase == "WAIT_PLUS35":
                state.plus35_pairs = pairs
                # update globals + persist
                globals()["GLOBAL_PLUS35_PAIRS"] = pairs
                save_pairs()

                bot.reply_to(
                    message,
                    f"✅ Збережено {len(pairs)} пар для 35% націнки. Можна продовжувати.",
                    parse_mode="HTML",
                )
                # phase stays as-is
                return

            return

        # === 2. Інші файли (BUDG, додаткові, Result) ===
        file_info = bot.get_file(message.document.file_id)
        file_bytes = bot.download_file(file_info.file_path)
        filename = message.document.file_name or "uploaded"

    except Exception as e:
        bot.reply_to(message, f"❌ Помилка читання файлу: <code>{e}</code>", parse_mode="HTML")
        return

    try:
        # --- ГОЛОВНА ТАБЛИЦЯ ---
        if state.phase == "WAIT_MAIN":
            df = load_main_budg_table(file_bytes, filename=filename)
            bot.reply_to(message, "✅ Головна таблиця завантажена! Тепер надішліть додаткові таблиці.")
            handle_main_table(message, state, df)
            return

        # --- ДОДАТКОВІ ТАБЛИЦІ ---
        elif state.phase == "WAIT_ADDITIONAL":
            df = read_additional_table(file_bytes, filename)
            handle_additional_table(message, state, df)
            return

        # --- РЕЗУЛЬТАТ (Result.xlsx) для алокації ---
        elif state.phase == "WAIT_ALLOC_RESULT":
            import io
            bio = io.BytesIO(file_bytes)
            try:
                df_res = pd.read_excel(bio, sheet_name="Result", engine="openpyxl")
            except Exception:
                bio.seek(0)
                df_res = pd.read_excel(bio, engine="openpyxl")

            # очищення назв колонок
            df_res.columns = [str(c).replace("\xa0", " ").replace("\u00A0", " ").strip() for c in df_res.columns]

            # базові числові колонки
            for num_col in ["FTD qty", "Total spend", "Total Dep Amount"]:
                if num_col in df_res.columns:
                    df_res[num_col] = pd.to_numeric(df_res[num_col], errors="coerce").fillna(0)

            state.alloc_df = df_res

            if state.alloc_mode == "alternative_leftover":
                # run allocation till L-adjust (no Pass#2), and get leftover
                out_df, leftover = allocate_total_spend_alternative(
                    df_res,
                    col_total_spend="Total spend",
                    col_ftd_qty="FTD qty",
                    col_cpa_target="CPA Target",
                    col_my_deposit="My deposit amount",
                    col_total_dep_amount="Total Dep Amount",
                    col_geo="ГЕО",
                    col_offer="Назва Офферу",
                    plus35_pairs=getattr(state, "plus35_pairs", None),
                    skip_pass2=True,
                    return_leftover=True,
                    excel_path="allocation_pass1_only.xlsx",
                    sheet_name="Result",
                )

                # send the file (optional)
                try:
                    with open("allocation_pass1_only.xlsx", "rb") as f:
                        bot.send_document(
                            message.chat.id,
                            f,
                            visible_file_name="allocation_pass1_only.xlsx")
                except Exception:
                    pass

                # send leftover summary
                bot.send_message(
                    message.chat.id,
                    f"Залишок бюджету після розподілу: $<b>{leftover:,.2f}</b>",
                    parse_mode="HTML"
                )

                state.phase = "WAIT_MAIN"
                return

            # якщо режим "alternative"
            if state.alloc_mode == "alternative":
                _run_alternative_and_send(chat_id, state)
                return

            # інші режими (openai або звичайна алокація)
            state.phase = "WAIT_ALLOC_BUDGET"
            bot.reply_to(message, "✅ Файл Result прийнято. Введіть, будь ласка, бюджет (наприклад: 200).")
            return

        # --- ОБ’ЄДНАННЯ UA/RU КРАЇН ДЛЯ /unite_geo ---
        elif state.phase == "WAIT_UNITE_TABLE":
            import io
            bio = io.BytesIO(file_bytes)
            fname = (filename or "").lower()
            try:
                if fname.endswith((".xlsx", ".xls", ".xlsm")):
                    df_in = pd.read_excel(bio)
                elif fname.endswith(".csv"):
                    bio.seek(0)
                    df_in = pd.read_csv(bio)
                else:
                    bot.reply_to(message, "⚠️ Надішліть файл у форматі .xlsx/.xls/.xlsm/.csv")
                    return
            except Exception as e:
                bot.reply_to(message, f"❌ Не вдалося прочитати файл: <code>{e}</code>", parse_mode="HTML")
                return

            # Виявити назви колонок
            country_col = next((c for c in ("ГЕО", "Гео", "GEO") if c in df_in.columns), None)
            if not country_col:
                bot.reply_to(message, "❌ У файлі немає колонки країни ('ГЕО' / 'Гео' / 'GEO').")
                return

            spend_col = next((c for c in ("Total spend", "Total Spend", "Загальні витрати") if c in df_in.columns),
                             None)
            if not spend_col:
                bot.reply_to(
                    message,
                    "❌ У файлі немає колонки витрат ('Total spend' / 'Total Spend' / 'Загальні витрати')."
                )
                return

            offer_col = "Назва Офферу"
            if offer_col not in df_in.columns:
                bot.reply_to(message, "❌ У файлі немає колонки 'Назва Офферу'.")
                return

            # 1) Об'єднати тільки в межах (Оффер + GEO UA/RU)
            try:
                merged = unite_ua_ru_rows(
                    df_in,
                    country_col=country_col,
                    total_spend_col=spend_col,
                    offer_col=offer_col,  # 👈 важливо!
                )
            except Exception as e:
                bot.reply_to(message, f"❌ Помилка під час об’єднання: <code>{e}</code>", parse_mode="HTML")
                return

            # Fix capitalization + special case for Congo
            if "ГЕО" in merged.columns:
                merged["ГЕО"] = merged["ГЕО"].apply(lambda x: str(x).capitalize() if isinstance(x, str) else x)
                merged["ГЕО"] = merged["ГЕО"].replace(r"(?i)^республіка\s+конго$", "Конго", regex=True)

            # 2) Побудувати вихід як повний Result і ПОВТОРНО застосувати формули
            #    (беремо всі потрібні колонки; якщо чогось немає — додаємо порожні)
            result_cols = [
                "Subid", "Offer ID", "Назва Офферу", "ГЕО",
                "FTD qty", "Total spend", "Total+%", "CPA", "CPA Target", "СP/Ч",
                "Total Dep Amount", "My deposit amount",
                "C. profit Target 40%", "C. profit Target 50%",
                "CAP", "Остаток CAP", "Current"
            ]

            # Привести імена в merged: перейменувати spend_col -> "Total spend" і GEO -> "ГЕО"
            merged = merged.copy()
            if spend_col != "Total spend" and spend_col in merged.columns:
                merged.rename(columns={spend_col: "Total spend"}, inplace=True)
            if country_col != "ГЕО" and country_col in merged.columns:
                merged.rename(columns={country_col: "ГЕО"}, inplace=True)

            # Забезпечити всі колонки
            for c in result_cols:
                if c not in merged.columns:
                    merged[c] = None

            # Привести порядок
            merged = merged[result_cols].copy()

            # Розрахунок Current (+/-) по збереженим парам
            st = user_states.get(message.chat.id)
            pairs = (getattr(st, "current_pairs", None) or globals().get("GLOBAL_CURRENT_PAIRS"))

            def _n(s: str) -> str:
                s = str(s or "").strip().lower()
                s = s.replace("’", "'").replace("`", "'").replace("–", "-").replace("—", "-")
                return " ".join(s.split())

            def _pair(r):
                return (_n(r.get("Назва Офферу", "")), _n(r.get("ГЕО", "")))

            merged["Current"] = merged.apply(lambda r: "+" if (pairs and _pair(r) in pairs) else "-", axis=1)

            # 3) Записати Excel + формули
            out = io.BytesIO()
            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                merged.to_excel(writer, index=False, sheet_name="Merged")
                ws = writer.book["Merged"]

                from openpyxl.styles import PatternFill, Alignment, Font
                from openpyxl.formatting.rule import FormulaRule

                header_row = 1
                first_row = 2
                last_row = ws.max_row

                # Множник за plus35_pairs
                p35 = (getattr(st, "plus35_pairs", None) or globals().get("GLOBAL_PLUS35_PAIRS"))

                # Формули
                for r in range(first_row, last_row + 1):
                    offer_val = ws[f"C{r}"].value  # Назва Офферу
                    geo_val = ws[f"D{r}"].value  # ГЕО

                    mul = 1.35 if (p35 and (_n(offer_val), _n(geo_val)) in p35) else 1.30

                    ws[f"G{r}"].value = f"=F{r}*{mul}"  # Total+%
                    ws[f"H{r}"].value = f"=IFERROR(G{r}/E{r},\"\")"  # CPA
                    ws[f"I{r}"].value = cpa_target_for_geo(ws[f"D{r}"].value)  # CPA Target
                    ws[f"J{r}"].value = f"=IFERROR(K{r}/E{r},\"\")"  # СP/Ч
                    ws[f"L{r}"].value = f"=IFERROR(K{r}/G{r}*100,0)"  # My deposit amount
                    ws[f"M{r}"].value = f"=G{r}*0.4"  # Profit 40%
                    ws[f"N{r}"].value = f"=G{r}*0.5"  # Profit 50%

                # Формати чисел
                for r in range(first_row, last_row + 1):
                    ws[f"E{r}"].number_format = "0"
                for c in ("F", "G", "H", "J", "K", "L", "M", "N"):
                    for r in range(first_row, last_row + 1):
                        ws[f"{c}{r}"].number_format = "0.00"

                # Шапка
                for col in range(1, 17):
                    ws.cell(row=1, column=col).font = Font(bold=True)
                    ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

                widths = {
                    "A": 10, "B": 12, "C": 22, "D": 16, "E": 10, "F": 14, "G": 12, "H": 10, "I": 12,
                    "J": 10, "K": 16, "L": 18, "M": 18, "N": 18, "O": 12, "P": 16
                }
                for col, w in widths.items():
                    ws.column_dimensions[col].width = w

                # (Опційно) твої ж умовні формати
                data_range = f"A{first_row}:P{last_row}"
                grey = PatternFill("solid", fgColor="BFBFBF")
                green = PatternFill("solid", fgColor="C6EFCE")
                yellow = PatternFill("solid", fgColor="FFEB9C")
                red = PatternFill("solid", fgColor="FFC7CE")

                ws.conditional_formatting.add(data_range, FormulaRule(formula=["$E2=0"], fill=grey, stopIfTrue=True))
                ws.conditional_formatting.add(
                    data_range,
                    FormulaRule(formula=["AND($E2>0,INT($H2)<=INT($I2),$L2>39)"], fill=green, stopIfTrue=True)
                )
                ws.conditional_formatting.add(
                    data_range,
                    FormulaRule(formula=["AND($E2>0,OR(INT($H2)<=INT($I2),AND($L2>39,$H2<$I2*1.31)))"], fill=yellow,
                                stopIfTrue=True)
                )
                ws.conditional_formatting.add(
                    data_range, FormulaRule(
                        formula=["OR(AND($E2>0,$H2>$I2*1.3,$L2>39),AND($E2>0,INT($H2)>INT($I2),$L2<39))"],
                        fill=red, stopIfTrue=True
                    )
                )

            out.seek(0)
            bot.send_document(
                chat_id,
                out,
                visible_file_name="merged_countries.xlsx",
                caption="✅ Готово! UA/RU країни об’єднано в межах одного оффера. Формули застосовано.",
            )

            state.phase = "WAIT_MAIN"
            return

        else:
            bot.reply_to(message, "⚠️ Несподівана фаза. Почніть спочатку з /start.")
            return

    except ValueError as ve:
        bot.reply_to(message, f"❌ Помилка у файлі <b>{filename}</b>:\n<code>{ve}</code>", parse_mode="HTML")
    except Exception as e:
        bot.reply_to(message, f"⚠️ Непередбачена помилка: <code>{e}</code>", parse_mode="HTML")


@bot.message_handler(commands=["allocate"])
@require_access
def cmd_allocate(message: types.Message):
    chat_id = message.chat.id
    state = user_states.setdefault(chat_id, UserState())

    # скинемо проміжний стан алокації
    state.alloc_df = None
    state.alloc_budget = None
    state.alloc_mode = None
    state.phase = "WAIT_ALLOC_MODE"

    kb = types.InlineKeyboardMarkup(row_width=2)
    # kb.add(
    #     types.InlineKeyboardButton("🔹 Прохід По Одному з KPI + Залишок", callback_data="alloc_mode:alternative"),
    # )
    kb.add(
        types.InlineKeyboardButton("🔹 Прохід По Одному з KPI", callback_data="alloc_mode:alternative_leftover"),
    )

    bot.reply_to(
        message,
        "Оберіть режим алокації:",
        reply_markup=kb
    )


@bot.callback_query_handler(func=lambda c: c.data and c.data.startswith("alloc_mode:"))
@require_access_cb
def on_alloc_mode(call: types.CallbackQuery):
    chat_id = call.message.chat.id
    state = user_states.setdefault(chat_id, UserState())
    mode = call.data.split(":", 1)[1]

    if mode not in {"optimal", "alternative", "alternative_leftover", "openai"}:
        bot.answer_callback_query(call.id, "Невідомий режим.")
        return

    state.alloc_mode = mode
    # Після вибору режиму просимо result.xlsx (для будь-якого режиму)
    state.phase = "WAIT_ALLOC_RESULT"
    bot.answer_callback_query(call.id, "Режим обрано.")
    bot.send_message(
        chat_id,
        (
            "Надішліть файл <b>result.xlsx</b> (той, що бот згенерував раніше).\n\n"
            "Після перерозподілу:\n"
            "- Спенд буде перерозподілений під прохід по 1 з КРІ.\n"
            "- Забереться спенд з офферів на яких відступні депозити та немає актуальних кап на поточний місяць\n"
            "- Бот виведе вам залишок спенда, після перерозподілу до одного з КРІ, який буде потрібно розкинути вручну."
        ),
        parse_mode="HTML"
    )


@bot.message_handler(content_types=["text"], func=lambda m: not (m.text or "").startswith("/"))
@require_access
def on_text(message: types.Message):
    chat_id = message.chat.id
    state = user_states.setdefault(chat_id, UserState())

    # перехоплюємо тільки у фазі алокації
    if state.phase != "WAIT_ALLOC_BUDGET":
        return

    # ===== Локальні режими: нижче все як було (потрібен бюджет) =====
    # Parse budget
    txt = (message.text or "").strip().replace(",", ".")
    try:
        budget = float(txt)
        if budget < 0:
            raise ValueError("negative")
    except Exception:
        bot.reply_to(message, "Введи, будь ласка, коректне додатнє число (наприклад: 200).")
        return

    if state.alloc_df is None or len(state.alloc_df) == 0:
        bot.reply_to(message, "Немає завантаженої таблиці Result. Використай /allocate ще раз.")
        state.phase = "WAIT_MAIN"
        return

    # Локальна алокація
    alloc_df, summary, alloc_vec = compute_optimal_allocation(state.alloc_df, budget)

    # Формуємо файл із новими витратами
    bio = io.BytesIO()
    write_result_like_excel_with_new_spend(bio, state.alloc_df, new_total_spend=alloc_vec)

    bio.seek(0)
    bot.send_document(
        chat_id,
        bio,
        visible_file_name="allocation.xlsx",
        caption=summary  # короткий підсумок
    )

    # Детальне пояснення
    explanation = build_allocation_explanation(state.alloc_df, alloc_vec, budget, max_lines=20)
    bot.send_message(chat_id, explanation)

    state.phase = "WAIT_MAIN"


@bot.callback_query_handler(func=lambda c: c.data == "skip_offer")
@require_access_cb
def on_skip_offer(call: types.CallbackQuery):
    chat_id = call.message.chat.id
    state = user_states.setdefault(chat_id, UserState())

    # Якщо вже поза межами — просто ігноруємо
    if state.current_offer_index >= len(state.offers):
        bot.answer_callback_query(call.id, "Немає активного оферу.")
        return

    offer = state.offers[state.current_offer_index]

    # ВАЖЛИВО: при пропуску — НЕ додаємо цей офер у результат,
    # тож просто прибираємо його з проміжних структур (якщо ти зберігаєш агрегати)
    if hasattr(state, "main_agg_df") and state.main_agg_df is not None:
        # повністю забираємо рядки цього оферу, щоб не потрапили у фінальний Excel
        state.main_agg_df = state.main_agg_df[state.main_agg_df["Назва Офферу"] != offer]

    # перехід до наступного оферу
    state.current_offer_index += 1
    bot.answer_callback_query(call.id, "Офер пропущено.")

    # якщо ще є офери — попросимо наступну додаткову таблицю
    if state.current_offer_index < len(state.offers):
        ask_additional_table_with_skip(call.message, state)
    else:
        # якщо оферів більше немає — генеруємо фінальний файл
        try:
            final_df = build_final_output(state)
            send_final_table(call.message, final_df)
        except Exception as e:
            bot.send_message(chat_id, f"⚠️ Помилка під час формування файлу: <code>{e}</code>")


@bot.message_handler(commands=["current"])
@require_access
def cmd_current(message: types.Message):
    st = user_states.setdefault(message.chat.id, UserState())
    st.phase = "WAIT_CURRENT"
    bot.reply_to(
        message,
        "📥 Надішліть файл <b>current.xlsx</b> або <b>current.csv</b> "
        "з колонками 'Назва Офферу' та 'ГЕО'. Цей файл буде збережено і не потрібно буде надсилати його щоразу.",
        parse_mode="HTML",
    )


@bot.message_handler(commands=["plus35"])
@require_access
def cmd_plus35(message: types.Message):
    st = user_states.setdefault(message.chat.id, UserState())
    st.phase = "WAIT_PLUS35"
    bot.reply_to(message, "Надішліть файл <b>plus35.xlsx</b>/<b>.csv</b> з колонками 'Назва Офферу' і 'ГЕО' (для 35%).",
                 parse_mode="HTML")


@bot.message_handler(commands=["whoami"])
def whoami(message: types.Message):
    bot.reply_to(message, f"Ваш Telegram ID: <code>{message.from_user.id}</code>")


@bot.message_handler(commands=["unite_geo"])
@require_access
def cmd_unite_geo(message: types.Message):
    st = user_states.setdefault(message.chat.id, UserState())
    st.phase = "WAIT_UNITE_TABLE"
    bot.reply_to(
        message,
        "📄 Надішліть Excel/CSV з колонками 'ГЕО' і 'Total spend' (або 'Загальні витрати'). "
        "Я об’єднаю UA/RU країни в один рядок з українською назвою і підсумком витрат.",
        parse_mode="HTML",
    )


# ===================== MAIN TABLE LOGIC =====================

def handle_main_table(message: types.Message, state: UserState, df: pd.DataFrame):
    # Clean & coerce
    work = df.copy()
    work["Назва Офферу"] = work["Назва Офферу"].astype(str).str.strip()
    work["ГЕО"] = work["ГЕО"].astype(str).str.strip()
    work["Загальні витрати"] = pd.to_numeric(work["Загальні витрати"], errors="coerce").fillna(0.0)

    # Drop empty/placeholder Offer IDs - handle string values properly
    work = work[
        work["Назва Офферу"].ne("") &
        work["Назва Офферу"].ne("nan") &
        work["Назва Офферу"].ne("None") &
        work["Назва Офферу"].notna()
        ]

    # Also filter out rows where ГЕО is empty
    work = work[
        work["ГЕО"].ne("") &
        work["ГЕО"].ne("nan") &
        work["ГЕО"].ne("None") &
        work["ГЕО"].notna()
        ]

    if len(work) == 0:
        bot.reply_to(message, "Не знайдено валідних записів у BUDG таблиці після очищення.")
        return

    # Aggregate by Offer ID + GEO
    agg = (
        work.groupby(["Назва Офферу", "ГЕО"])["Загальні витрати"]
        .sum().reset_index()
    )

    state.main_agg_df = agg

    # Unique Offer IDs (from cleaned data)
    state.offers = sorted(agg["Назва Офферу"].unique().tolist())
    state.phase = "WAIT_ADDITIONAL"
    state.current_offer_index = 0
    ask_additional_table_with_skip(message, state)

    if not state.offers:
        bot.reply_to(message, "Не знайдено жодного валідного Offer ID у аркуші BUDG після очищення.")
        return


# ===================== ADDITIONAL TABLE LOGIC =====================

def handle_additional_table(message: types.Message, state: UserState, df: pd.DataFrame):
    # 1) Clean & normalize
    work = df.copy()
    work["Країна"] = work["Країна"].astype(str).str.strip()
    work["Сума депозитів"] = pd.to_numeric(work["Сума депозитів"], errors="coerce").fillna(0.0)

    # Фільтруємо порожні/некоректні країни
    work = work[
        work["Країна"].ne("") &
        work["Країна"].ne("nan") &
        work["Країна"].ne("None") &
        work["Країна"].notna()
        ]

    # 2) Визначаємо поточний офер
    try:
        current_offer = state.offers[state.current_offer_index]
    except IndexError:
        bot.reply_to(message, "Помилка: немає активного Offer ID. Напиши /start для початку.")
        return

    # 3) Якщо після очищення немає жодного валідного рядка — проставляємо нулі
    if len(work) == 0:
        # Зберігаємо «нульові» депозити для логіки фінального мерджу
        # (порожній словник означає, що по країнах нічого не додавати;
        # нижче ще й гарантуємо нулі у проміжній таблиці, якщо вона вже є)
        state.offer_deposits[current_offer] = {}

        # Якщо в пам’яті вже є агрегована головна таблиця — гарантуємо нулі для поточного оферу
        if hasattr(state, "main_agg_df") and state.main_agg_df is not None:
            mask = state.main_agg_df["Назва Офферу"] == current_offer
            # створимо колонки, якщо їх ще немає
            if "Total Dep Sum" not in state.main_agg_df.columns:
                state.main_agg_df["Total Dep Sum"] = 0.0
            if "Total Dep Amount" not in state.main_agg_df.columns:
                state.main_agg_df["Total Dep Amount"] = 0
            # нулі для всіх рядків цього оферу
            state.main_agg_df.loc[mask, "Total Dep Sum"] = 0.0
            state.main_agg_df.loc[mask, "Total Dep Amount"] = 0

        # Повідомлення користувачу
        bot.reply_to(
            message,
            (
                f"ℹ️ У додатковій таблиці для <b>{current_offer}</b> немає даних після очищення.\n"
                f"Для цього оферу проставлено <b>0</b> у колонках депозитів."
            ),
        )

        # Перехід до наступного оферу / фінал
        state.current_offer_index += 1
        if state.current_offer_index >= len(state.offers):
            final_df = build_final_output(state)
            send_final_table(message, final_df)
            user_states[message.chat.id] = UserState()  # reset
            return

        next_offer = state.offers[state.current_offer_index]
        bot.reply_to(
            message,
            (
                f"Надішліть додаткову таблицю для <b>{next_offer}</b> "
                f"({state.current_offer_index + 1}/{len(state.offers)})."
            ),
        )
        return

    # 4) Якщо дані є — канонікалізуємо країни та агрегуємо
    work["canon_en"] = work["Країна"].apply(
        lambda x: to_canonical_en(x, state.country_map_uk_to_en, state.country_canon, state.country_map_ru_to_en)
    )

    dep_by_country = (
        work.groupby("canon_en")["Сума депозитів"]
        .agg(["sum", "count"]).reset_index()
        .rename(columns={"sum": "total", "count": "count"})
    )

    # Зберігаємо агрегати в пам'ять для цього оферу
    state.offer_deposits[current_offer] = {
        row["canon_en"]: {"total": float(row["total"]), "count": int(row["count"])}
        for _, row in dep_by_country.iterrows()
    }

    # 5) Підсумкова інфо для користувача
    countries_found = list(dep_by_country["canon_en"].unique())
    total_deposits = float(dep_by_country["total"].sum())

    # 6) Перехід до наступного оферу / фінал
    state.current_offer_index += 1
    if state.current_offer_index >= len(state.offers):
        final_df = build_final_output(state)
        send_final_table(message, final_df)
        user_states[message.chat.id] = UserState()  # reset
        return

    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton("Пропустити цей офер", callback_data="skip_offer"))

    next_offer = state.offers[state.current_offer_index]
    summary = f"""
✅ Прийнято дані для <b>{current_offer}</b>
📊 Знайдено {len(countries_found)} країн, загальна сума депозитів: {total_deposits:,.2f}

Країни: {', '.join(countries_found[:5])}{' ...' if len(countries_found) > 5 else ''}

Надішли наступну додаткову таблицю для <b>{next_offer}</b> ({state.current_offer_index + 1}/{len(state.offers)})
    """.strip()

    bot.send_message(message.from_user.id, summary, reply_markup=kb)


# ===================== BUILD FINAL =====================

def geo_to_canonical(
        geo: str,
        uk_to_en: Dict[str, str],
        canonical: Dict[str, str],
        ru_to_en: Optional[Dict[str, str]] = None,
) -> str:
    return to_canonical_en(geo, uk_to_en, canonical, ru_to_en)


def build_final_output(state: UserState) -> pd.DataFrame:
    agg = state.main_agg_df.copy()
    # Canonical GEO for matching
    agg["ГЕО_canon"] = agg["ГЕО"].apply(
        lambda g: geo_to_canonical(g, state.country_map_uk_to_en, state.country_canon, state.country_map_ru_to_en))

    rows: List[Dict[str, object]] = []
    for _, row in agg.iterrows():
        offer_name = str(row["Назва Офферу"])
        # If you have a real Offer ID elsewhere — put it here. Fallback to offer_name.
        offer_id = offer_name

        geo_display = str(row["ГЕО"])
        geo_canon = str(row["ГЕО_canon"])
        spend_total = float(row["Загальні витрати"])

        dep_sum = 0.0  # Total Dep Sum ($)
        dep_cnt = 0  # FTD qty
        offer_map = state.offer_deposits.get(offer_name, {})
        if geo_canon in offer_map:
            dep_sum = float(offer_map[geo_canon]["total"])
            dep_cnt = int(offer_map[geo_canon]["count"])
        else:
            import difflib
            candidates = list(offer_map.keys())
            close = difflib.get_close_matches(geo_canon, candidates, n=1, cutoff=0.6)
            if close:
                dep_sum = float(offer_map[close[0]]["total"])
                dep_cnt = int(offer_map[close[0]]["count"])

        rows.append({
            "Subid": "",  # empty
            "Offer ID": offer_id,  # from main table (fallback = name)
            "Назва Офферу": offer_name,  # Offer Name
            "ГЕО": geo_display,  # Country
            "FTD qty": dep_cnt,  # count
            "Total Spend": spend_total,  # $
            "Total Dep Amount": dep_sum,  # $ (your naming; this is the sum)
            # rest computed later in Excel
        })

    # Order primary columns; computed ones will be added in send_final_table
    df = pd.DataFrame(rows, columns=[
        "Subid", "Offer ID", "Назва Офферу", "ГЕО", "FTD qty", "Total Spend", "Total Dep Amount"
    ])
    return df


# ===================== SENDER =====================

def send_final_table(message: types.Message, df: pd.DataFrame):
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        # Rebuild to requested order
        base_cols = [
            "Subid", "Offer ID", "Назва Офферу", "ГЕО",
            "FTD qty", "Total Spend", "Total Dep Amount"
        ]
        df = df[base_cols].copy()

        st = user_states.get(message.chat.id)

        # ✅ фолбек на глобальні пари
        pairs = (getattr(st, "current_pairs", None) or GLOBAL_CURRENT_PAIRS)
        p35 = (getattr(st, "plus35_pairs", None) or GLOBAL_PLUS35_PAIRS)

        # (не обов'язково) локальний нормалізатор
        def _n(s: str) -> str:
            s = str(s or "").strip().lower()
            s = s.replace("’", "'").replace("`", "'").replace("–", "-").replace("—", "-")
            return " ".join(s.split())

        def _pair(row):
            return (_n(row.get("Назва Офферу", "")), _n(row.get("ГЕО", "")))

        # Колонка Current
        if pairs:
            df["Current"] = df.apply(lambda r: "+" if _pair(r) in pairs else "-", axis=1)
        else:
            df["Current"] = "-"

        # ---- NEW: round numeric inputs before write ----
        df["Total Spend"] = pd.to_numeric(df["Total Spend"], errors="coerce").round(2)
        df["Total Dep Amount"] = pd.to_numeric(df["Total Dep Amount"], errors="coerce").round(2)
        df["FTD qty"] = pd.to_numeric(df["FTD qty"], errors="coerce").fillna(0).astype(int)
        # -------------------------------------------------

        final_cols = [
            "Subid", "Offer ID", "Назва Офферу", "ГЕО",
            "FTD qty", "Total spend", "Total+%", "CPA", "CPA Target", "СP/Ч",
            "Total Dep Amount", "My deposit amount", "C. profit Target 40%", "C. profit Target 50%",
            "CAP", "Остаток CAP", "Current"
        ]

        df.rename(columns={"Total Spend": "Total spend"}, inplace=True)

        # Placeholders
        df["Total+%"] = None
        df["CPA"] = None
        df["CPA Target"] = None
        df["СP/Ч"] = None
        df["My deposit amount"] = None
        df["C. profit Target 40%"] = None
        df["C. profit Target 50%"] = None
        df["CAP"] = ""
        df["Остаток CAP"] = ""

        df = df[final_cols]

        df.to_excel(writer, index=False, sheet_name="Result")

        wb = writer.book
        ws = writer.sheets["Result"]

        from openpyxl.styles import PatternFill, Alignment, Font
        from openpyxl.formatting.rule import FormulaRule

        first_row = 2
        last_row = ws.max_row

        # === Формули ===
        # G: Total+% — ТЕПЕР з урахуванням plus35_pairs
        for r in range(first_row, last_row + 1):
            offer_val = ws[f"C{r}"].value  # Назва Офферу
            geo_val = ws[f"D{r}"].value  # ГЕО

            mul = 1.35 if (p35 and (_n(offer_val), _n(geo_val)) in p35) else 1.3
            ws[f"G{r}"].value = f"=F{r}*{mul}"  # Total+%

            ws[f"H{r}"].value = f"=IFERROR(G{r}/E{r},\"\")"  # CPA
            ws[f"I{r}"].value = cpa_target_for_geo(ws[f"D{r}"].value)  # CPA Target
            ws[f"J{r}"].value = f"=IFERROR(K{r}/E{r},\"\")"  # СP/Ч
            ws[f"L{r}"].value = f"=IFERROR(K{r}/G{r}*100,0)"  # My deposit amount
            ws[f"M{r}"].value = f"=G{r}*0.4"  # Profit 40%
            ws[f"N{r}"].value = f"=G{r}*0.5"  # Profit 50%

        # ---- Формати чисел ----
        # Integers: E (FTD qty)
        for r in range(first_row, last_row + 1):
            ws[f"E{r}"].number_format = "0"

        # Two decimals: F..N (окрім I — там ціле значення цілі)
        two_dec_cols = ["F", "G", "H", "J", "K", "L", "M", "N"]
        for col in two_dec_cols:
            for r in range(first_row, last_row + 1):
                ws[f"{col}{r}"].number_format = "0.00"

        # --- Header styling ---
        for col in range(1, 17):  # A..P
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

        widths = {
            "A": 10, "B": 12, "C": 22, "D": 16, "E": 10, "F": 14, "G": 12, "H": 10, "I": 12,
            "J": 10, "K": 16, "L": 18, "M": 18, "N": 18, "O": 12, "P": 16
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # --- Conditional formatting ---
        data_range = f"A{first_row}:P{last_row}"
        grey = PatternFill("solid", fgColor="BFBFBF")
        green = PatternFill("solid", fgColor="C6EFCE")
        yellow = PatternFill("solid", fgColor="FFEB9C")
        red = PatternFill("solid", fgColor="FFC7CE")

        # Grey: E = 0
        ws.conditional_formatting.add(
            data_range, FormulaRule(formula=["$E2=0"], fill=grey, stopIfTrue=True)
        )
        # Green: INT(H) <= INT(I) AND L > 39
        ws.conditional_formatting.add(
            data_range, FormulaRule(formula=["AND($E2>0,INT($H2)<=INT($I2),$L2>39)"], fill=green, stopIfTrue=True)
        )
        # Yellow: (INT(H) <= INT(I)) OR (L > 39 AND H < I*1.31)
        ws.conditional_formatting.add(
            data_range, FormulaRule(formula=["AND($E2>0,OR(INT($H2)<=INT($I2),AND($L2>39,$H2<$I2*1.31)))"], fill=yellow,
                                    stopIfTrue=True)
        )
        # Red:
        ws.conditional_formatting.add(
            data_range, FormulaRule(
                formula=["OR(AND($E2>0,$H2>$I2*1.3,$L2>39),AND($E2>0,INT($H2)>INT($I2),$L2<39))"],
                fill=red, stopIfTrue=True
            )
        )

    bio.seek(0)
    bot.send_document(
        message.chat.id,
        bio,
        visible_file_name="result.xlsx",
        caption="Фінальна таблиця з готовим аналізом 📊"
    )


# ===================== MAIN =====================
if __name__ == "__main__":
    print("Bot is running...")
    bot.infinity_polling(skip_pending=True)
