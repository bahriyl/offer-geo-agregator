import importlib
import io
import os
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

os.environ.setdefault("BOT_TOKEN", "0:TEST")

from main import (  # noqa: E402
    EPS_YEL,
    RED_MULT,
    _build_threshold_table,
    _extract_targets,
    compute_allocation_max_yellow,
)


def test_allocation_continues_to_red_cap_when_budget_remains():
    df = pd.DataFrame(
        {
            "FTD qty": [10, 12],
            "Total spend": [200.0, 200.0],
            "Total Dep Amount": [150.0, 180.0],
            "Total+%": [500.0, 500.0],
            "CPA Target": [8.0, 8.0],
        },
        index=["row_a", "row_b"],
    )

    result_df, used_budget, alloc_vec = compute_allocation_max_yellow(df)
    assert not result_df.empty

    E = pd.to_numeric(df.get("FTD qty"), errors="coerce").fillna(0.0)
    K = pd.to_numeric(df.get("Total Dep Amount"), errors="coerce").fillna(0.0)
    targets, target_ints = _extract_targets(df)
    thresholds = _build_threshold_table(E, K, targets, target_ints)

    red_caps = thresholds["red_ceiling"].astype(float)
    yellow_caps = thresholds["yellow_soft_ceiling"].astype(float)

    expected_red_caps = (
        (targets.to_numpy(dtype=float) * RED_MULT * E.to_numpy(dtype=float)) / 1.3
    ) - EPS_YEL
    expected_red_caps = np.maximum(expected_red_caps, 0.0)

    np.testing.assert_allclose(red_caps.to_numpy(dtype=float), expected_red_caps, rtol=1e-6, atol=1e-6)

    assert float(df["Total spend"].sum()) > float(yellow_caps.sum()) + 1e-6

    np.testing.assert_allclose(alloc_vec.to_numpy(dtype=float), red_caps.to_numpy(dtype=float))
    assert used_budget == pytest.approx(float(red_caps.sum()))
    assert used_budget >= float(yellow_caps.sum())


def _prep_allocation_inputs(main_mod, df):
    dfw = df.copy()
    dfw.columns = [str(c).strip() for c in dfw.columns]
    E = pd.to_numeric(dfw.get("FTD qty", 0.0), errors="coerce").fillna(0.0)
    F_original = main_mod._normalize_money(
        dfw.get("Total spend", pd.Series(0.0, index=dfw.index))
    ).fillna(0.0)
    K = main_mod._normalize_money(dfw.get("Total Dep Amount", pd.Series(0.0, index=dfw.index))).fillna(0.0)
    T = pd.to_numeric(dfw.get("Total+%", 0.0), errors="coerce").fillna(0.0)
    targets, target_ints = main_mod._extract_targets(dfw)
    thresholds = main_mod._build_threshold_table(E, K, targets, target_ints)

    stop_before_red = thresholds["red_ceiling"].fillna(0.0)
    row_allowance = pd.Series(
        np.minimum(T.to_numpy(), stop_before_red.to_numpy()),
        index=dfw.index,
    ).clip(lower=0.0)

    partner_series = dfw.get(
        "Partner",
        dfw.get("Offer ID", dfw.get("Назва Офферу", pd.Series([""] * len(dfw), index=dfw.index))),
    )
    partner_series = partner_series.fillna("").astype(str)
    geo_series = dfw.get("ГЕО", pd.Series([""] * len(dfw), index=dfw.index)).fillna("").astype(str)

    order_df = pd.DataFrame(
        {
            "original_spend": F_original,
            "partner": partner_series,
            "geo": geo_series,
            "__pos": np.arange(len(dfw), dtype=int),
        },
        index=dfw.index,
    )

    spend_order = order_df.sort_values(
        by=["original_spend", "partner", "geo", "__pos"],
        ascending=[True, True, True, True],
        kind="mergesort",
    ).index.tolist()

    return {
        "E": E,
        "F_original": F_original,
        "K": K,
        "T": T,
        "targets": targets,
        "thresholds": thresholds,
        "row_allowance": row_allowance,
        "yellow_caps": thresholds["yellow_soft_ceiling"].fillna(0.0).clip(lower=0.0),
        "red_caps": stop_before_red,
        "available_budget": float(F_original.sum()),
        "order": spend_order,
    }


def test_low_spend_row_receives_leftover_before_high_spend():
    os.environ["BOT_TOKEN"] = "123:ABC"
    main_mod = importlib.reload(importlib.import_module("main"))

    df = pd.DataFrame(
        {
            "FTD qty": [40, 80],
            "Total spend": [50.0, 200.0],
            "Total Dep Amount": [20.0, 78.0],
            "Total+%": [80.0, 360.0],
            "CPA Target": [10.0, 20.0],
        }
    )

    inputs = _prep_allocation_inputs(main_mod, df)
    order = inputs["order"]
    row_allowance = inputs["row_allowance"]
    yellow_caps = inputs["yellow_caps"]
    red_caps = inputs["red_caps"].clip(lower=0.0)
    F_original = inputs["F_original"]
    rem = inputs["available_budget"]

    alloc_expected = pd.Series(0.0, index=df.index, dtype=float)

    for idx in order:
        if rem <= 1e-9:
            break
        if float(inputs["E"].at[idx]) <= 0:
            continue
        cap = min(float(yellow_caps.at[idx]), float(row_allowance.at[idx]))
        cap = max(cap, 0.0)
        need = cap - float(alloc_expected.at[idx])
        if need <= 1e-9:
            continue
        give = min(rem, need)
        alloc_expected.at[idx] += give
        rem -= give

    if rem > 1e-9:
        for idx in order:
            if rem <= 1e-9:
                break
            if float(inputs["E"].at[idx]) <= 0:
                continue
            cap = min(float(red_caps.at[idx]), float(row_allowance.at[idx]))
            cap = max(cap, 0.0)
            need = cap - float(alloc_expected.at[idx])
            if need <= 1e-9:
                continue
            give = min(rem, need)
            alloc_expected.at[idx] += give
            rem -= give

    result_df, used_budget, alloc_vec = main_mod.compute_allocation_max_yellow(df)

    np.testing.assert_allclose(
        alloc_vec.to_numpy(dtype=float),
        alloc_expected.to_numpy(dtype=float),
        rtol=1e-9,
        atol=1e-9,
    )
    assert used_budget == pytest.approx(float(alloc_expected.sum()))

    low_idx = F_original.sort_values(ascending=True).index[0]
    high_idx = F_original.sort_values(ascending=True).index[-1]
    expected_low_cap = min(float(red_caps.at[low_idx]), float(row_allowance.at[low_idx]))
    assert alloc_vec.at[low_idx] == pytest.approx(expected_low_cap)

    assert result_df.loc[low_idx, "Allocated extra"] == pytest.approx(alloc_vec.at[low_idx])


def test_allocation_parses_currency_strings_with_non_standard_formats():
    os.environ["BOT_TOKEN"] = "789:XYZ"
    main_mod = importlib.reload(importlib.import_module("main"))

    df = pd.DataFrame(
        {
            "FTD qty": [20, 15, 12],
            "Total spend": ["1\u00a0234,56", "1,234.56", "$500"],
            "Total Dep Amount": ["800", "650", "$300"],
            "Total+%": [2500, 2200, 1200],
            "CPA Target": [8.0, 9.0, 10.0],
        },
        index=["nbsp_comma", "comma_dot", "currency"],
    )

    result_df, used_budget, alloc_vec = main_mod.compute_allocation_max_yellow(df)

    expected = np.array([1234.56, 1234.56, 500.0])
    parsed = main_mod._normalize_money(df.get("Total spend")).to_numpy(dtype=float)
    np.testing.assert_allclose(parsed, expected, rtol=1e-9)

    assert used_budget > 0.0
    assert used_budget == pytest.approx(float(alloc_vec.sum()))
    assert (alloc_vec.to_numpy(dtype=float) > 0.0).all()
    assert alloc_vec.index.tolist() == df.index.tolist()


def test_classify_status_marks_red_only_above_cutoff_and_excel_rule_matches():
    os.environ["BOT_TOKEN"] = "147:STATUS"
    main_mod = importlib.reload(importlib.import_module("main"))

    e = 13.0
    target = 8.0
    red_cutoff = target * main_mod.RED_MULT

    cpa_yellow_edge = target * main_mod.YELLOW_MULT - 0.05
    f_yellow_edge = (cpa_yellow_edge * e) / 1.3
    deposit_yellow_edge = 1.3 * f_yellow_edge * 0.5
    assert main_mod._classify_status(e, f_yellow_edge, deposit_yellow_edge, target) == "Yellow"

    cpa_lower = target * main_mod.YELLOW_MULT
    f_lower = (cpa_lower * e) / 1.3
    deposit_lower = 1.3 * f_lower * 0.5
    status_lower = main_mod._classify_status(e, f_lower, deposit_lower, target)
    assert status_lower == "Red"

    cpa_mid = target * (main_mod.YELLOW_MULT + main_mod.RED_MULT) / 2
    f_mid = (cpa_mid * e) / 1.3
    deposit_mid = 1.3 * f_mid * 0.5
    status_mid = main_mod._classify_status(e, f_mid, deposit_mid, target)
    assert status_mid == "Red"

    cpa_equal = red_cutoff
    f_equal = (cpa_equal * e) / 1.3
    deposit_equal = 1.3 * f_equal * 0.5
    status_equal = main_mod._classify_status(e, f_equal, deposit_equal, target)
    assert status_equal == "Red"

    cpa_far = red_cutoff + 1.0
    f_far = (cpa_far * e) / 1.3
    deposit_far = 1.3 * f_far * 0.5
    assert main_mod._classify_status(e, f_far, deposit_far, target) == "Grey"

    df = pd.DataFrame(
        {
            "Subid": ["s_yellow", "s_lower", "s_mid", "s_upper", "s_far"],
            "Offer ID": ["o1", "o2", "o3", "o4", "o5"],
            "Назва Офферу": ["Offer", "Offer", "Offer", "Offer", "Offer"],
            "ГЕО": ["G1", "G2", "G3", "G4", "G5"],
            "FTD qty": [e, e, e, e, e],
            "Total spend": [0.0, 0.0, 0.0, 0.0, 0.0],
            "Total Dep Amount": [
                deposit_yellow_edge,
                deposit_lower,
                deposit_mid,
                deposit_equal,
                deposit_far,
            ],
            "CPA Target": [target, target, target, target, target],
        },
        index=["s_yellow", "s_lower", "s_mid", "s_upper", "s_far"],
    )

    bio = io.BytesIO()
    new_spend = pd.Series([f_yellow_edge, f_lower, f_mid, f_equal, f_far], index=df.index)
    main_mod.write_result_like_excel_with_new_spend(
        bio,
        df,
        new_spend,
        overwrite_total_spend=True,
    )
    bio.seek(0)
    wb = load_workbook(bio)
    ws = wb["Result"]

    upper_row_excel = df.index.get_loc("s_upper") + 2
    lower_row_excel = df.index.get_loc("s_lower") + 2

    e_cell = ws[f"E{upper_row_excel}"]
    f_cell = ws[f"F{upper_row_excel}"]
    i_cell = ws[f"I{upper_row_excel}"]

    assert e_cell.value == pytest.approx(e)
    assert f_cell.value == pytest.approx(f_equal, rel=0, abs=1e-6)

    computed_cpa = 1.3 * f_cell.value / e_cell.value
    assert computed_cpa == pytest.approx(red_cutoff)
    assert computed_cpa >= i_cell.value * main_mod.YELLOW_MULT - main_mod.CPA_TOL
    assert computed_cpa <= i_cell.value * main_mod.RED_MULT + main_mod.CPA_TOL

    lower_cpa = 1.3 * ws[f"F{lower_row_excel}"].value / ws[f"E{lower_row_excel}"].value
    assert lower_cpa == pytest.approx(target * main_mod.YELLOW_MULT)

    red_rule_formulae = []
    for rules in ws.conditional_formatting._cf_rules.values():
        for rule in rules:
            if getattr(rule, "type", None) != "expression":
                continue
            formulas = getattr(rule, "formula", [])
            if isinstance(formulas, str):
                formulas = [formulas]
            for formula in formulas:
                if "$H2>$I2" in formula or "$H2>=$I2" in formula:
                    red_rule_formulae.append(formula)

    assert any(
        (f"$H2>=$I2*{main_mod.YELLOW_MULT:.2f}" in f)
        and (f"$H2<=$I2*{main_mod.RED_MULT:.2f}" in f)
        for f in red_rule_formulae
    )


def test_allocation_explanation_reflects_custom_targets_in_status_counts():
    os.environ["BOT_TOKEN"] = "258:CUSTOM"
    main_mod = importlib.reload(importlib.import_module("main"))

    df = pd.DataFrame(
        {
            "Назва Офферу": ["Offer X"],
            "ГЕО": ["UA"],
            "FTD qty": [11],
            "Total spend": [50.0],
            "Total Dep Amount": [400.0],
            "Total+%": [120.0],
            "CPA Target": [5.0],
        }
    )

    alloc_vec = pd.Series([65.0], index=df.index, dtype=float)

    explanation = main_mod.build_allocation_explanation(
        df,
        alloc_vec,
        budget=float(df["Total spend"].sum()),
        alloc_is_delta=False,
    )

    assert "Жовтих ДО/ПІСЛЯ: 1 → 0" in explanation
    assert "Yellow → Red" in explanation


def test_read_result_allocation_table_handles_formula_total_plus_percent(tmp_path):
    os.environ["BOT_TOKEN"] = "456:FORM"
    main_mod = importlib.reload(importlib.import_module("main"))

    source_df = pd.DataFrame(
        {
            "FTD qty": [25],
            "Total spend": [180.0],
            "Total Dep Amount": [120.0],
            "Total+%": ["=B2*1.3"],
            "CPA Target": [9.0],
        }
    )

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        source_df.to_excel(writer, sheet_name="Result", index=False)
    file_bytes = bio.getvalue()

    parsed = main_mod.read_result_allocation_table(file_bytes, "result.xlsx")
    assert pytest.approx(parsed.at[0, "Total spend"], rel=1e-9) == 180.0
    assert parsed.at[0, "Total+%"] > parsed.at[0, "Total spend"]

    E = pd.to_numeric(parsed.get("FTD qty"), errors="coerce").fillna(0.0)
    F = parsed["Total spend"].astype(float)
    K = parsed["Total Dep Amount"].astype(float)
    targets, target_ints = main_mod._extract_targets(parsed)
    thresholds = main_mod._build_threshold_table(E, K, targets, target_ints)
    stop_before_red = thresholds["red_ceiling"].fillna(0.0)
    row_allowance = pd.Series(
        np.minimum(parsed["Total+%"].to_numpy(dtype=float), stop_before_red.to_numpy(dtype=float)),
        index=parsed.index,
    ).clip(lower=0.0)

    assert row_allowance.iloc[0] > 0.0

    result_df, used_budget, alloc_vec = main_mod.compute_allocation_max_yellow(parsed)
    assert used_budget > 0.0
    assert alloc_vec.iloc[0] > 0.0
    assert result_df.loc[parsed.index[0], "Allocated extra"] >= 0.0


def test_yellow_formula_written_to_excel_matches_helper():
    os.environ["BOT_TOKEN"] = "456:FORMULA"
    main_mod = importlib.reload(importlib.import_module("main"))

    df = pd.DataFrame(
        {
            "Subid": ["s1"],
            "Offer ID": ["o1"],
            "Назва Офферу": ["Offer"],
            "ГЕО": ["Geo"],
            "FTD qty": [5],
            "Total spend": [100.0],
            "Total Dep Amount": [50.0],
        }
    )

    bio = io.BytesIO()
    main_mod.write_result_like_excel_with_new_spend(
        bio,
        df,
        pd.Series([0.0]),
        overwrite_total_spend=True,
    )
    bio.seek(0)
    wb = load_workbook(bio)
    ws = wb["Result"]

    expected_formula = main_mod._build_yellow_formula()
    formulas = []
    for rules in ws.conditional_formatting._cf_rules.values():
        for rule in rules:
            if getattr(rule, "type", None) != "expression":
                continue
            formula_field = getattr(rule, "formula", None)
            if not formula_field:
                continue
            if isinstance(formula_field, (list, tuple)):
                formulas.extend(formula_field)
            else:
                formulas.append(formula_field)

    assert expected_formula in formulas
